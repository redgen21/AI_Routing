"""Pure data contract for the shared Region Plan v2 Excel workflow.

This module deliberately creates no SQL and writes no files.  The API owns durable
storage and activation; it receives the deterministic artifact returned here.
"""
from __future__ import annotations

import hashlib, io, json, re
from collections import Counter
from datetime import date
from pathlib import PurePosixPath
from typing import Any, Mapping

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None

SCHEMA_VERSION = "region-plan-workflow/v2"
IMPORTER_VERSION = "region-plan-v2-data/1"
AREA_COLUMNS = ("region_code", "region_name", "region_seq", "postal_code", "area_type", "required_center_type", "membership_rank", "is_primary", "overflow_allowed", "overflow_penalty_minutes", "overflow_reason")
TECHNICIAN_COLUMNS = ("technician_id", "region_code", "active", "policy_mode", "effective_from", "effective_to")
POLICY_MODES = frozenset((
    "home_distance_only", "preferred_region_soft",
    "assigned_region_boundary_spillover", "active_roster_type_hard_region_soft",
    "active_roster_area_type_fallback_region_soft",
))
POLICY_VERSION_BY_MODE = {
    "home_distance_only": "home_distance_only",
    "preferred_region_soft": "preferred_region_soft",
    "assigned_region_boundary_spillover": "explicit_workbook_membership/v1",
    "active_roster_type_hard_region_soft": "active_roster_type_hard_region_soft/v1",
    "active_roster_area_type_fallback_region_soft": "active_roster_area_type_fallback_region_soft/v1",
}
POLICY_VERSIONS_BY_MODE = {
    "home_distance_only": frozenset({"home_distance_only"}),
    "preferred_region_soft": frozenset({"preferred_region_soft"}),
    "assigned_region_boundary_spillover": frozenset({
        "explicit_workbook_membership/v1",
        "own_region_with_approved_boundary_overflow/v2",
    }),
    "active_roster_type_hard_region_soft": frozenset({"active_roster_type_hard_region_soft/v1"}),
    "active_roster_area_type_fallback_region_soft": frozenset({"active_roster_area_type_fallback_region_soft/v1"}),
}
_ALIASES = {
    "zipcode":"postal_code", "zip code":"postal_code", "territory":"region_code",
    "region id":"region_code", "region seq":"region_seq", "area name":"region_name",
    "new region name":"region_name", "area type":"area_type",
    "required center type":"required_center_type", "membership rank":"membership_rank",
    "is primary":"is_primary", "overflow allowed":"overflow_allowed",
    "overflow penalty minutes":"overflow_penalty_minutes", "overflow reason":"overflow_reason",
    "tech id":"technician_id", "assignment":"region_code", "tech name":"_ignored_pii",
}
_SAFE_CITY = re.compile(r"^[A-Za-z0-9][A-Za-z0-9_ .,&()/-]{0,159}$")

class RegionPlanV2ValidationError(ValueError):
    def __init__(self, code: str): self.code=code; super().__init__(code)

def _key(value: Any) -> str: return re.sub(r"[ _-]+", " ", str(value or "").strip().lower())
def _value(value: Any) -> str: return "" if value is None else str(value).strip()
def _sha(data: bytes) -> str: return hashlib.sha256(data).hexdigest()
def _bool(value: str) -> bool:
    if value.lower() in ("true","1","yes"): return True
    if value.lower() in ("false","0","no"): return False
    raise ValueError

def _legacy_region_code(value: str) -> str:
    """Preserve supplied stable IDs and compact only display-only legacy labels."""
    supplied = value.strip()
    if re.fullmatch(r"[A-Za-z0-9][A-Za-z0-9_.-]{0,79}", supplied):
        return supplied
    token=re.sub(r"[^A-Za-z0-9]", "", supplied).upper()
    if re.fullmatch(r"[A-Z0-9_-]{1,4}", token): return token
    return "R" + _sha(value.strip().casefold().encode())[:3].upper()

def state_layout(state_root: str, environment: str, subsidiary_id: str, target_city_id: str, workbook_sha256: str) -> dict[str, str]:
    """Return allowlisted relative artifact locations; caller performs all I/O."""
    for part in (environment, subsidiary_id, target_city_id, workbook_sha256):
        if not re.fullmatch(r"[A-Za-z0-9_-]+", part): raise RegionPlanV2ValidationError("STATE_PATH_INVALID")
    base = PurePosixPath(state_root) / "v2" / environment / subsidiary_id / target_city_id / workbook_sha256
    return {name: str(base / name) for name in ("source.xlsx", "source.sha256", "canonical.json", "validation.json", "rejects.jsonl", "receipt.json")}

def _sheet_rows(sheet, expected: tuple[str,...], sheet_name: str):
    rows=list(sheet.iter_rows(values_only=True)); physical=max(0,len(rows)-1)
    if not rows: return [], physical, []
    headers=[]; seen=set()
    for h in rows[0]:
        raw=_value(h); canonical=_ALIASES.get(_key(raw), _key(raw).replace(" ","_"))
        if canonical in seen: raise RegionPlanV2ValidationError("CANONICAL_COLUMN_INVALID")
        seen.add(canonical); headers.append(canonical)
    unknown=[h for h in headers if h not in expected and h != "_ignored_pii" and not h.startswith("extension.")]
    area_required=("region_code","postal_code","area_type")
    tech_required=("technician_id","region_code")
    if unknown or any(x not in headers for x in (area_required if sheet_name=="Area" else tech_required)):
        raise RegionPlanV2ValidationError("CANONICAL_COLUMN_INVALID")
    output=[]; rejects=[]
    for n,row in enumerate(rows[1:],2):
        vals={headers[i]:_value(v) for i,v in enumerate(row) if i<len(headers)}
        if not any(vals.values()): continue
        output.append((n,vals))
    return output, physical, rejects

def canonicalize_workbook(workbook_bytes: bytes, city_metadata: Mapping[str,Any]) -> dict[str,Any]:
    """Validate an uploaded .xlsx and return an immutable, JSON-safe candidate payload."""
    city_metadata = dict(city_metadata)
    city_name = _value(city_metadata.get("city_name") or city_metadata.get("source_city_id") or city_metadata.get("target_city_id"))
    if city_name:
        city_metadata["city_name"] = city_name
        city_metadata.setdefault("source_city_id", city_name)
        city_metadata.setdefault("target_city_id", city_name)
    required=("subsidiary_id","target_city_id","source_city_id")
    if any(not _value(city_metadata.get(k)) for k in required): raise RegionPlanV2ValidationError("CITY_METADATA_MISSING")
    if not re.fullmatch(r"[A-Za-z0-9_-]{1,80}", _value(city_metadata["subsidiary_id"])) or not re.fullmatch(r"[A-Za-z0-9][A-Za-z0-9_ .,&()/-]{0,159}", city_name):
        raise RegionPlanV2ValidationError("CITY_METADATA_INVALID")
    selected_mode=_value(city_metadata.get("technician_policy_mode"))
    selected_policy=_value(city_metadata.get("policy_version"))
    if selected_mode or selected_policy:
        if selected_mode not in POLICY_MODES or selected_policy not in POLICY_VERSIONS_BY_MODE[selected_mode]:
            raise RegionPlanV2ValidationError("CITY_METADATA_MISMATCH")
    if not workbook_bytes.startswith(b"PK") or load_workbook is None: raise RegionPlanV2ValidationError("WORKBOOK_FORMAT_INVALID")
    try: wb=load_workbook(io.BytesIO(workbook_bytes), read_only=True, data_only=True)
    except Exception as e: raise RegionPlanV2ValidationError("WORKBOOK_FORMAT_INVALID") from e
    sheet_aliases={"Area":"Area","1. Area":"Area","Technician":"Technician","2. Technician":"Technician"}
    canonical_sheets={}
    for name in wb.sheetnames:
        canonical=sheet_aliases.get(name)
        if canonical is None: raise RegionPlanV2ValidationError("REQUIRED_SHEET_MISSING")
        if canonical in canonical_sheets: raise RegionPlanV2ValidationError("AMBIGUOUS_SHEET_ALIAS")
        canonical_sheets[canonical]=wb[name]
    if set(canonical_sheets) != {"Area","Technician"}: raise RegionPlanV2ValidationError("REQUIRED_SHEET_MISSING")
    areas,ap,_=_sheet_rows(canonical_sheets["Area"], AREA_COLUMNS,"Area"); techs,tp,_=_sheet_rows(canonical_sheets["Technician"], TECHNICIAN_COLUMNS,"Technician")
    rejects=[]; plan_errors=[]; normalized_areas=[]; regions={}
    for row, x in areas:
        try:
            x={k:_value(x.get(k)) for k in AREA_COLUMNS}
            # Territory is a legacy label, not a v2 identifier.  Preserve it only as a
            # display name and derive a short deterministic identifier.
            if not x["region_name"]: x["region_name"]=x["region_code"]
            x["region_code"]=_legacy_region_code(x["region_code"]); x["postal_code"]=x["postal_code"].zfill(5); x["area_type"]=x["area_type"].upper(); x["required_center_type"]=(x["required_center_type"] or x["area_type"]).upper()
            if x["region_seq"]:
                x["region_seq"] = str(int(float(x["region_seq"])))
            if not re.fullmatch(r"[A-Za-z0-9][A-Za-z0-9_.-]{0,79}",x["region_code"]) or not re.fullmatch(r"\d{5}",x["postal_code"]) or x["area_type"] not in ("DMS","DMS2") or x["required_center_type"] != x["area_type"]: raise ValueError
            x["membership_rank"]=int(x["membership_rank"] or 1); x["is_primary"]=_bool(x["is_primary"] or "true"); x["overflow_allowed"]=_bool(x["overflow_allowed"] or "false")
            penalty=x["overflow_penalty_minutes"]
            if x["overflow_allowed"]:
                parsed_penalty=float(penalty)
                if parsed_penalty <= 0 or not parsed_penalty.is_integer(): raise ValueError
                x["overflow_penalty_minutes"]=int(parsed_penalty)
            else:
                if penalty: raise ValueError
                x["overflow_penalty_minutes"]=None
            if x["region_code"] in regions and regions[x["region_code"]][:3] != (x["region_name"],x["area_type"],x["required_center_type"]): raise ValueError
            if x["region_code"] in regions and x["region_seq"] and regions[x["region_code"]][3] not in {None, x["region_seq"]}: raise ValueError
            regions[x["region_code"]]=(x["region_name"],x["area_type"],x["required_center_type"],x["region_seq"] or None); normalized_areas.append(x)
        except Exception: rejects.append({"sheet":"Area","row_number":row,"error_code":"AREA_ROW_INVALID"})
    explicit_sequences = {value[3] for value in regions.values() if value[3] is not None}
    if explicit_sequences and len(explicit_sequences) != len(regions):
        plan_errors.append("REGION_SEQ_INCOMPLETE")
    if explicit_sequences:
        try:
            sequence_values = sorted(int(value) for value in explicit_sequences)
        except ValueError:
            sequence_values = []
        if sequence_values != list(range(1, len(regions) + 1)):
            plan_errors.append("REGION_SEQ_INVALID")
        sequence_by_code = {code: int(value[3]) for code, value in regions.items() if value[3] is not None}
    else:
        sequence_by_code = {code: index for index, code in enumerate(sorted(regions), 1)}
    for item in normalized_areas:
        item["region_seq"] = str(sequence_by_code.get(item["region_code"], ""))
    postal=Counter(a["postal_code"] for a in normalized_areas)
    for p,count in postal.items():
        members=[a for a in normalized_areas if a["postal_code"]==p]
        if sum(a["is_primary"] for a in members)!=1 or sorted(a["membership_rank"] for a in members)!=list(range(1,count+1)):
            plan_errors.append("OVERLAP_POLICY_INVALID")
        if count == 1 and (members[0]["overflow_allowed"] or members[0]["overflow_penalty_minutes"] is not None or members[0]["overflow_reason"]):
            plan_errors.append("OVERLAP_POLICY_INVALID")
        if count > 1 and any(not a["is_primary"] and not a["overflow_allowed"] for a in members):
            plan_errors.append("OVERLAP_POLICY_INVALID")
    normalized_techs=[]; assigned=set()
    for row,x in techs:
        try:
            x={k:_value(x.get(k)) for k in TECHNICIAN_COLUMNS}; x["technician_id"]=x["technician_id"].upper(); x["region_code"]=_legacy_region_code(x["region_code"]) if x["region_code"] else ""
            if x["region_code"] not in regions:
                name_matches = [code for code, value in regions.items() if value[0] == x["region_code"]]
                if len(name_matches) == 1:
                    x["region_code"] = name_matches[0]
            if not x["technician_id"]: raise ValueError
            if not x["region_code"]: raise RegionPlanV2ValidationError("TECHNICIAN_ASSIGNMENT_BLANK")
            if x["technician_id"] in assigned or x["region_code"] not in regions: raise ValueError
            assigned.add(x["technician_id"]); x["active"]=_bool(x["active"] or "true")
            # The selected Plan policy is authoritative.  Area Map workbooks
            # may have been generated under an earlier policy; technician
            # membership is reusable and must not make a new Plan impossible
            # to create under a different, explicitly selected policy.
            x["policy_mode"]=selected_mode
            start=date.fromisoformat(x["effective_from"]) if x["effective_from"] else None
            end=date.fromisoformat(x["effective_to"]) if x["effective_to"] else None
            if start and end and start>end: raise ValueError
            normalized_techs.append(x)
        except RegionPlanV2ValidationError as e: rejects.append({"sheet":"Technician","row_number":row,"error_code":e.code})
        except Exception: rejects.append({"sheet":"Technician","row_number":row,"error_code":"TECHNICIAN_ROW_INVALID"})
    accounting={"Area":{"physical_rows":ap,"input_rows":len(areas),"accepted_rows":len(normalized_areas),"rejected_rows":sum(r['sheet']=='Area' for r in rejects)},"Technician":{"physical_rows":tp,"input_rows":len(techs),"accepted_rows":len(normalized_techs),"rejected_rows":sum(r['sheet']=='Technician' for r in rejects)}}
    if not any(t["active"] for t in normalized_techs): plan_errors.append("NO_ACTIVE_TECHNICIAN")
    metadata = dict(city_metadata)
    # A policy city can host several Area Plans (for example Atlanta_GA hosts
    # Atlanta_3area and Atlanta_6area).  Keep the Area Plan name independent
    # from target_city_id and make it part of the human-readable immutable ID.
    # The checksum suffix still makes every changed candidate distinct.
    display_name = _value(metadata.get("plan_display_name"))
    plan_key = re.sub(r"[^A-Za-z0-9]+", "_", display_name).strip("_").lower()
    if not plan_key:
        plan_key = re.sub(r"[^A-Za-z0-9]+", "_", _value(metadata["target_city_id"])).strip("_").lower()
    if not plan_key:
        raise RegionPlanV2ValidationError("PLAN_DISPLAY_NAME_INVALID")
    metadata["plan_display_name"] = display_name or _value(metadata["target_city_id"])
    manifest={"schema_version":SCHEMA_VERSION,"importer_version":IMPORTER_VERSION,"city_metadata":metadata,"source_workbook_sha256":_sha(workbook_bytes),"areas":sorted(normalized_areas,key=lambda x:(x['postal_code'],x['region_code'])),"technicians":sorted(normalized_techs,key=lambda x:x['technician_id']),"row_accounting":accounting,"plan_errors":sorted(set(plan_errors))}
    fatal_rejects=[r for r in rejects if r["error_code"] != "TECHNICIAN_ASSIGNMENT_BLANK"]
    manifest["excluded_rows"]={"TECHNICIAN_ASSIGNMENT_BLANK":sum(r["error_code"]=="TECHNICIAN_ASSIGNMENT_BLANK" for r in rejects)}
    canonical_bytes=json.dumps(manifest,sort_keys=True,separators=(",",":"),ensure_ascii=False).encode(); canonical_sha=_sha(canonical_bytes); manifest["canonical_sha256"]=canonical_sha; manifest["plan_id"]=f"rp2_{plan_key[:80]}_{canonical_sha[:20]}"; manifest["status"]="rejected" if fatal_rejects or plan_errors else "candidate"
    return {"manifest":manifest,"rejects":rejects,"artifacts":{"canonical.json":canonical_bytes,"rejects.jsonl":b"".join(json.dumps(r,sort_keys=True).encode()+b"\n" for r in rejects)}}

def adopt_legacy_candidate(metadata: Mapping[str,Any], *, expected_area_count: int, expected_technician_count: int, expected_filenames: tuple[str,...]) -> dict[str,Any]:
    """Adopt a caller-identified six-file bundle after strict manifest/count checks."""
    files=tuple(sorted(str(x) for x in (metadata.get("files") or metadata.get("bundle") or [])))
    if files != tuple(sorted(expected_filenames)) or len(files)!=6 or int(metadata.get("area_count",-1))!=expected_area_count or int(metadata.get("technician_count",-1))!=expected_technician_count: raise RegionPlanV2ValidationError("LEGACY_BUNDLE_INVALID")
    if not all(re.fullmatch(r"[0-9a-f]{64}", str(metadata.get(k,""))) for k in ("manifest_sha256","source_sha256")): raise RegionPlanV2ValidationError("LEGACY_MANIFEST_INVALID")
    payload={"schema_version":SCHEMA_VERSION,"legacy_adapter":"six-file/v1","source_workbook_sha256":metadata["source_sha256"],"legacy_manifest_sha256":metadata["manifest_sha256"],"area_count":expected_area_count,"technician_count":expected_technician_count,"files":list(files)}
    digest=_sha(json.dumps(payload,sort_keys=True,separators=(",",":")).encode()); payload["canonical_sha256"]=digest; return payload

def adopt_legacy_la_candidate(metadata: Mapping[str,Any]) -> dict[str,Any]:
    """Compatibility wrapper; city identity belongs to the API caller, never this module."""
    names=tuple(metadata.get("expected_filenames") or ())
    return adopt_legacy_candidate(metadata, expected_area_count=413, expected_technician_count=54, expected_filenames=names)

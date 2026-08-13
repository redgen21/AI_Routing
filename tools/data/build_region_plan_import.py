"""Build deterministic, development-only generic region-plan import bundles.

This converter deliberately performs no database work.  It converts the two
canonical workbook sheets into PII-redacted CSVs, a lineage/quality manifest,
and idempotent PostgreSQL upserts for the additive V001 region-plan schema.
"""
from __future__ import annotations

import argparse, csv, hashlib, json, re
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

SCHEMA = "region-workbook-import/v1"
POLICY_EXPLICIT = "explicit_workbook_membership/v1"
POLICY_ATLANTA_CURRENT = "own_region_with_approved_boundary_overflow/v2"
POLICY_TYPE_HARD_REGION_SOFT = "active_roster_type_hard_region_soft/v1"
POLICY_AREA_TYPE_FALLBACK_REGION_SOFT = "active_roster_area_type_fallback_region_soft/v1"
POLICY_MODES = {
    "home_distance_only": "home_distance_only",
    "preferred_region_soft": "preferred_region_soft",
    POLICY_EXPLICIT: "assigned_region_boundary_spillover",
    POLICY_ATLANTA_CURRENT: "assigned_region_boundary_spillover",
    POLICY_TYPE_HARD_REGION_SOFT: "active_roster_type_hard_region_soft",
    POLICY_AREA_TYPE_FALLBACK_REGION_SOFT: "active_roster_area_type_fallback_region_soft",
}
AREA_CLASSIFIED_POLICIES = frozenset(
    {POLICY_TYPE_HARD_REGION_SOFT, POLICY_AREA_TYPE_FALLBACK_REGION_SOFT}
)
# Employee codes are source-system identifiers, not an Atlanta-specific AI######
# value.  Keep the accepted alphabet deliberately narrow so generated SQL and
# CSV artifacts remain deterministic while supporting LA's numeric/alphabetic
# roster IDs.
TECH_RE = re.compile(r"^[A-Za-z0-9][A-Za-z0-9._-]{0,63}$")
POSTAL_RE = re.compile(r"^\d{5}$")

def sha(p): return hashlib.sha256(p.read_bytes()).hexdigest()
def clean(v): return "" if v is None else str(v).strip()
def postal(v):
    s = clean(v)
    if isinstance(v, float) and v.is_integer(): s = str(int(v))
    if isinstance(v, int): s = str(v)
    return s.zfill(5) if s.isdigit() and len(s) <= 5 else s
def csv_write(path, fields, rows):
    with path.open("w", newline="", encoding="utf-8") as f:
        w=csv.DictWriter(f, fieldnames=fields, lineterminator="\n"); w.writeheader(); w.writerows(rows)
def esc(v): return "'" + str(v).replace("'", "''") + "'"
def relation(value, territories):
    """Parse supplied territory labels around explicit > / <> evidence."""
    for op in ("<>", ">"):
        if op in value:
            left,right=(x.strip() for x in value.split(op,1))
            # A numeric token is an explicit caller-supplied stable region
            # identifier (1-based sequence), not a Zone-specific assumption.
            names=list(territories)
            if left.isdigit() and 1 <= int(left) <= len(names): left=names[int(left)-1]
            if right.isdigit() and 1 <= int(right) <= len(names): right=names[int(right)-1]
            if left in territories and right in territories: return left,op,right
    return None

def build(source: Path, out_root: Path, *, city=None, plan=None, territory_order=None,
          source_technician_city=None, policy_version=None, overflow_enabled=True):
    """Generic workbook conversion; scenario defaults below are CLI-only."""
    if not city or not plan or not territory_order or not source_technician_city or not policy_version:
        raise ValueError("city, plan, territories, source_technician_city and policy_version are required")
    if policy_version not in POLICY_MODES:
        raise ValueError("POLICY_VERSION_INVALID")
    if policy_version in AREA_CLASSIFIED_POLICIES and overflow_enabled:
        raise ValueError("AREA_TYPE_REGION_SOFT_REQUIRES_DISABLED_OVERFLOW")
    policy_mode = POLICY_MODES[policy_version]
    if city == source_technician_city:
        raise ValueError("TARGET_CITY_MUST_DIFFER_FROM_SOURCE_TECHNICIAN_CITY")
    seq={x:i+1 for i,x in enumerate(territory_order)}
    wb=openpyxl.load_workbook(source, data_only=True, read_only=True)
    expected_area=("ZIPCode","Territory","Area Type")
    expected_tech=("Tech ID","Tech Name","Assignment")
    if tuple(clean(x) for x in next(wb["1. Area"].iter_rows(min_row=1,max_row=1,values_only=True))) != expected_area:
        raise ValueError("AREA_HEADER_CONTRACT_INVALID")
    if tuple(clean(x) for x in next(wb["2. Technician"].iter_rows(min_row=1,max_row=1,values_only=True))) != expected_tech:
        raise ValueError("TECHNICIAN_HEADER_CONTRACT_INVALID")
    area=list(wb["1. Area"].iter_rows(min_row=2, values_only=True)); tech=list(wb["2. Technician"].iter_rows(min_row=2, values_only=True))
    rejects=[]; memberships=[]
    for n,row in enumerate(area,2):
        z,t,kind=(list(row)+[None,None,None])[:3]; z,t,kind=postal(z),clean(t),clean(kind)
        if not (z or t or kind): continue # blank formatting tails are not input rows
        m=relation(t,seq) if overflow_enabled else None
        if not POSTAL_RE.fullmatch(z): rejects.append({"sheet":"1. Area","source_row":n,"reason":"INVALID_POSTAL"}); continue
        if kind not in ("DMS","DMS2"): rejects.append({"sheet":"1. Area","source_row":n,"reason":"INVALID_AREA_TYPE"}); continue
        if m:
            primary, relation_syntax, alternate = m
            memberships.append((z,primary,kind,n,relation_syntax,alternate))
        elif t in seq: memberships.append((z,t,kind,n,"", ""))
        else: rejects.append({"sheet":"1. Area","source_row":n,"reason":"UNKNOWN_TERRITORY"})
    tech_rows=[]; seen=set()
    for n,row in enumerate(tech,2):
        code,name,t=(list(row)+[None,None,None])[:3]; code,t=clean(code),clean(t)
        if not (code or clean(name) or t): continue
        if not t:
            rejects.append({"sheet":"2. Technician","source_row":n,"reason":"BLANK_TECHNICIAN_ASSIGNMENT"}); continue
        if not TECH_RE.fullmatch(code): rejects.append({"sheet":"2. Technician","source_row":n,"reason":"INVALID_TECHNICIAN_ID"}); continue
        if code in seen: rejects.append({"sheet":"2. Technician","source_row":n,"reason":"DUPLICATE_TECHNICIAN_ID"}); continue
        if t not in seq: rejects.append({"sheet":"2. Technician","source_row":n,"reason":"UNKNOWN_TECHNICIAN_ASSIGNMENT"}); continue
        seen.add(code); tech_rows.append({"employee_code":code,"assigned_region_seq":seq[t],"assigned_territory":t,"policy_mode":policy_mode,"source_row":n})
    grouped=defaultdict(list)
    for x in memberships: grouped[x[0]].append(x)
    postal_rows=[]; overflow=[]
    for z, rows in sorted(grouped.items()):
        # duplicate exact ordinary source rows are rejected; overlap syntax makes two memberships explicit.
        first=rows[0]; all_regions=[first[1]] + ([first[5]] if first[5] else [])
        distinct_later=[r for r in rows[1:] if r[1] != first[1]]
        if overflow_enabled and distinct_later and not first[5]:
            # A repeated ZIP with a different explicit territory is source
            # evidence of alternate membership; first source row is primary.
            all_regions += [r[1] for r in distinct_later]
        if len(rows)>1 and first[5]:
            # Retain any distinct later territory as another explicit alternate;
            # only exact ZIP+territory repeats are rejects.
            for r in rows[1:]:
                if r[1] == first[1] or r[1] == first[5]:
                    rejects.append({"sheet":"1. Area","source_row":r[3],"reason":"DUPLICATE_POSTAL_SOURCE_ROW"})
                else:
                    all_regions.append(r[1]); distinct_later.append(r)
        elif len(rows)>1 and not (overflow_enabled and distinct_later):
            # Keep the earliest valid occurrence as the stable canonical value;
            # each later physical input row is explicitly rejected.
            rejects.extend({"sheet":"1. Area","source_row":r[3],"reason":"DUPLICATE_POSTAL_SOURCE_ROW"} for r in rows[1:])
        postal_rows.append({"postal_code":z,"primary_region_seq":seq[first[1]],"primary_territory":first[1],"area_type":first[2],"source_membership_count":len(all_regions),"source_region_seqs":json.dumps([seq[x] for x in all_regions],separators=(",",":")),"source_row":first[3]})
        alternates=list(dict.fromkeys(([first[5]] if first[5] else []) + [r[1] for r in distinct_later]))
        if len(all_regions) > 2:
            raise ValueError(f"POSTAL_MORE_THAN_TWO_MEMBERSHIPS:{z}")
        for alternate in alternates:
            overflow.append({"postal_code":z,"primary_region_seq":seq[first[1]],"alternate_region_seq":seq[alternate],"relation_syntax":first[4] or "duplicate_territory","allow_overflow":"true","penalty_cost":4500,"rationale":"explicit workbook syntax: "+first[1]+(" "+first[4]+" " if first[4] else " duplicate territory ")+alternate})
    postal_counts=Counter(r["primary_territory"] for r in postal_rows)
    empty=[name for name in territory_order if not postal_counts[name]]
    if empty: raise ValueError("DECLARED_REGION_WITHOUT_POSTAL:"+",".join(empty))
    technician_counts=Counter(r["assigned_territory"] for r in tech_rows)
    if not tech_rows: raise ValueError("NO_VALID_TECHNICIANS")
    territory_area_types={name:sorted({r["area_type"] for r in postal_rows if r["primary_territory"]==name}) for name in territory_order}
    required_center_types={name:(types[0] if len(types)==1 else None) for name,types in territory_area_types.items()}
    if policy_version in AREA_CLASSIFIED_POLICIES:
        mixed=[name for name,types in territory_area_types.items() if len(types)!=1]
        if mixed: raise ValueError("TERRITORY_AREA_TYPE_NOT_UNIFORM:"+",".join(mixed))
    requires_source_id_migration=any(not re.fullmatch(r"AI\d{6}",r["employee_code"]) for r in tech_rows)
    out=out_root/city; out.mkdir(parents=True,exist_ok=True)
    csv_write(out/"region_postals.csv", list(postal_rows[0]) if postal_rows else [], postal_rows)
    csv_write(out/"technician_assignments.csv", list(tech_rows[0]) if tech_rows else [], tech_rows)
    csv_write(out/"boundary_overflow.csv", ["postal_code","primary_region_seq","alternate_region_seq","relation_syntax","allow_overflow","penalty_cost","rationale"], overflow)
    csv_write(out/"rejects.csv", ["sheet","source_row","reason"], rejects)
    source_sha=sha(source); files={p.name:sha(p) for p in out.glob("*.csv")}
    area_input=len([r for r in area if any(x is not None and clean(x) for x in r)])
    area_rejected=len([r for r in rejects if r['sheet']=='1. Area'])
    unstaffed=[name for name in territory_order if not technician_counts[name]]
    required_migrations=["V001__atlanta_6area_region_plan"] + (["V003__region_plan_technician_source_id"] if requires_source_id_migration else []) + (["V004__region_plan_area_type_region_soft"] if policy_version in AREA_CLASSIFIED_POLICIES else [])
    manifest={"schema":"region-workbook-import/v1","environment":"development","write_target":"vrp_db_dev only","production_write_prohibited":True,"lifecycle_stage":"candidate","review_activation_required":True,"candidate_import_mutates_runtime_masters":False,"required_migrations":required_migrations,"source":{"path":str(source),"sha256":source_sha,"parent_workbook":source.name,"technician_source_city":source_technician_city},"plan_id":plan,"city":city,"policy_version":policy_version,"technician_policy_mode":policy_mode,"territories":[{"seq":v,"source_territory":k,"required_center_type":required_center_types[k],"area_type_evidence":territory_area_types[k],"canonical_postal_count":postal_counts[k],"technician_count":technician_counts[k]} for k,v in seq.items()],"regions_without_technicians":unstaffed,"row_accounting":{"area_input_nonblank":area_input,"area_accepted":area_input-area_rejected,"area_rejected":area_rejected,"canonical_unique_postals":len(postal_rows),"technician_input":len(tech),"technician_accepted":len(tech_rows),"technician_rejected":len([r for r in rejects if r['sheet']=='2. Technician'])},"null_policy":"blank formatting rows ignored; populated invalid rows rejected","duplicate_policy":"exact duplicate keeps earliest row and rejects later rows; distinct-territory duplicate creates primary plus overflow only when enabled","overflow_policy":{"enabled":overflow_enabled,"explicit_relation_syntax_accepted":overflow_enabled,"distinct_territory_duplicate_accepted":overflow_enabled},"technician_name_policy":"source names excluded from generated artifacts; employee_code is master lookup key","units":{"postal_code":"US ZIP five-digit string","region_seq":"integer","penalty_cost":"policy cost units"},"artifacts":files}
    manifest_path=out/"manifest.json"; manifest_path.write_text(json.dumps(manifest,indent=2)+"\n",encoding="utf-8")
    manifest_sha=sha(manifest_path); fixed_sha=files["region_postals.csv"]; boundary_sha=files["boundary_overflow.csv"]; tech_sha=files["technician_assignments.csv"]
    bundle_sha=hashlib.sha256("".join(files[k] for k in sorted(files)).encode()).hexdigest()
    # SQL is idempotent and guarded: it must be run only after selecting vrp_db_dev.
    sql=["-- DEVELOPMENT ONLY. Refuse any non-development database.","BEGIN TRANSACTION ISOLATION LEVEL SERIALIZABLE;","DO $$ BEGIN IF current_database() <> 'vrp_db_dev' THEN RAISE EXCEPTION 'refusing target %; expected vrp_db_dev', current_database(); END IF; IF "+esc(city)+" = "+esc(source_technician_city)+" THEN RAISE EXCEPTION 'target and source technician city must differ'; END IF; PERFORM pg_advisory_xact_lock(hashtext('LGEAI:'||"+esc(city)+")); END $$;"]
    sql += [f"DO $$ BEGIN IF NOT EXISTS (SELECT 1 FROM common_city_context WHERE subsidiary_name='LGEAI' AND strategic_city_name={esc(city)}) AND (EXISTS (SELECT 1 FROM common_region_master WHERE subsidiary_name='LGEAI' AND strategic_city_name={esc(city)}) OR EXISTS (SELECT 1 FROM common_technician_master WHERE subsidiary_name='LGEAI' AND strategic_city_name={esc(city)}) OR EXISTS (SELECT 1 FROM common_technician_capability_master WHERE subsidiary_name='LGEAI' AND strategic_city_name={esc(city)})) THEN RAISE EXCEPTION 'target city has unmanaged existing data'; END IF; END $$;", f"INSERT INTO common_city_context (subsidiary_name,strategic_city_name,source_strategic_city_name,context_version,policy_version) VALUES ('LGEAI',{esc(city)},{esc(source_technician_city)},'v1',{esc(policy_version)}) ON CONFLICT (subsidiary_name,strategic_city_name) DO NOTHING;", f"DO $$ BEGIN IF EXISTS (SELECT 1 FROM common_city_context WHERE subsidiary_name='LGEAI' AND strategic_city_name={esc(city)} AND source_strategic_city_name <> {esc(source_technician_city)}) THEN RAISE EXCEPTION 'existing city source technician context differs'; END IF; IF EXISTS (SELECT 1 FROM common_region_plan WHERE subsidiary_name='LGEAI' AND strategic_city_name={esc(city)} AND plan_id={esc(plan)} AND (source_sha256 <> {esc(source_sha)} OR policy_version <> {esc(policy_version)})) THEN RAISE EXCEPTION 'existing plan_id has different source checksum or policy'; END IF; IF EXISTS (SELECT 1 FROM common_region_plan WHERE subsidiary_name='LGEAI' AND strategic_city_name={esc(city)} AND plan_id={esc(plan)} AND bundle_sha256 IS NOT NULL AND bundle_sha256 <> {esc(bundle_sha)}) THEN RAISE EXCEPTION 'existing plan_id has different canonical bundle'; END IF; END $$;"]
    if requires_source_id_migration: sql += ["DO $$ BEGIN IF NOT EXISTS (SELECT 1 FROM pg_constraint WHERE conrelid='public.common_region_plan_technician'::regclass AND conname='common_region_plan_technician_employee_code_source_id_check') THEN RAISE EXCEPTION 'V003__region_plan_technician_source_id is required'; END IF; END $$;"]
    if policy_version in AREA_CLASSIFIED_POLICIES: sql += ["DO $$ BEGIN IF NOT EXISTS (SELECT 1 FROM information_schema.columns WHERE table_schema='public' AND table_name='common_region_plan_region' AND column_name='required_center_type') OR NOT EXISTS (SELECT 1 FROM pg_constraint WHERE conrelid='public.common_region_plan_technician'::regclass AND conname='common_region_plan_technician_policy_mode_v004_check') THEN RAISE EXCEPTION 'V004__region_plan_area_type_region_soft is required'; END IF; END $$;"]
    if len(seq)>6: sql += ["DO $$ BEGIN IF EXISTS (SELECT 1 FROM pg_constraint WHERE conname='common_region_plan_region_region_seq_check') OR NOT EXISTS (SELECT 1 FROM pg_constraint WHERE conname='common_region_plan_region_region_seq_positive_check') THEN RAISE EXCEPTION 'V002 unbounded region-sequence migration required'; END IF; END $$;"]
    sql += [f"INSERT INTO common_region_plan (subsidiary_name,strategic_city_name,plan_id,schema_version,policy_version,source_file_name,source_sha256,manifest_sha256,bundle_sha256,fixed_region_sha256,boundary_policy_sha256,technician_policy_sha256,membership_input_rows,membership_accepted_rows,membership_rejected_rows,unique_postal_count,technician_count,ambiguous_postal_count,import_idempotency_key,imported_by) VALUES ('LGEAI',{esc(city)},{esc(plan)},'{SCHEMA}',{esc(policy_version)},{esc(source.name)},{esc(source_sha)},{esc(manifest_sha)},{esc(bundle_sha)},{esc(fixed_sha)},{esc(boundary_sha)},{esc(tech_sha)},{manifest['row_accounting']['area_input_nonblank']},{manifest['row_accounting']['area_accepted']},{manifest['row_accounting']['area_rejected']},{len(postal_rows)},{len(tech_rows)},{len(overflow)},{esc(plan+':'+policy_version+':'+source_sha)},'generated_converter') ON CONFLICT (subsidiary_name,strategic_city_name,plan_id) DO NOTHING;"]
    ids=','.join(esc(r['employee_code']) for r in tech_rows)
    sql += [f"DO $$ BEGIN IF (SELECT count(*) FROM common_technician_master WHERE subsidiary_name='LGEAI' AND strategic_city_name={esc(source_technician_city)} AND employee_code IN ({ids}) AND active_flag) <> {len(tech_rows)} THEN RAISE EXCEPTION 'source technician roster missing/inactive'; END IF; END $$;"]
    for name,n in seq.items():
        base=f"'LGEAI',{esc(city)},{esc(plan)},{n},{esc(plan+'_r'+str(n).zfill(2))},{esc(city+' '+name)},{esc(name)}"
        if policy_version in AREA_CLASSIFIED_POLICIES:
            sql.append(f"INSERT INTO common_region_plan_region (subsidiary_name,strategic_city_name,plan_id,region_seq,region_id,region_name,source_territory,required_center_type) VALUES ({base},{esc(required_center_types[name])}) ON CONFLICT DO NOTHING;")
        else:
            sql.append(f"INSERT INTO common_region_plan_region (subsidiary_name,strategic_city_name,plan_id,region_seq,region_id,region_name,source_territory) VALUES ({base}) ON CONFLICT DO NOTHING;")
    for r in postal_rows:
        status='not_required' if r['source_membership_count']==1 else 'resolved'
        sql.append(f"INSERT INTO common_region_plan_postal (subsidiary_name,strategic_city_name,plan_id,postal_code,region_seq,area_type,source_membership_count,resolution_status,source_region_seqs) VALUES ('LGEAI',{esc(city)},{esc(plan)},{esc(r['postal_code'])},{r['primary_region_seq']},{esc(r['area_type'])},{r['source_membership_count']},{esc(status)},{esc(r['source_region_seqs'])}::jsonb) ON CONFLICT DO NOTHING;")
    for r in tech_rows: sql.append(f"INSERT INTO common_region_plan_technician (subsidiary_name,strategic_city_name,plan_id,employee_code,assigned_region_seq,policy_mode) VALUES ('LGEAI',{esc(city)},{esc(plan)},{esc(r['employee_code'])},{r['assigned_region_seq']},{esc(policy_mode)}) ON CONFLICT DO NOTHING;")
    for r in overflow: sql.append(f"INSERT INTO common_region_plan_boundary_overflow (subsidiary_name,strategic_city_name,plan_id,postal_code,primary_region_seq,alternate_region_seq,allow_overflow,penalty_cost,rationale,policy_version) VALUES ('LGEAI',{esc(city)},{esc(plan)},{esc(r['postal_code'])},{r['primary_region_seq']},{r['alternate_region_seq']},true,4500,{esc(r['rationale'])},{esc(policy_version)}) ON CONFLICT DO NOTHING;")
    if policy_version in AREA_CLASSIFIED_POLICIES:
        region_columns="region_seq,region_id,region_name,source_territory,required_center_type"
        region_values=",".join(f"({n},{esc(plan+'_r'+str(n).zfill(2))},{esc(city+' '+name)},{esc(name)},{esc(required_center_types[name])})" for name,n in seq.items())
    else:
        region_columns="region_seq,region_id,region_name,source_territory"
        region_values=",".join(f"({n},{esc(plan+'_r'+str(n).zfill(2))},{esc(city+' '+name)},{esc(name)})" for name,n in seq.items())
    postal_values=",".join(f"({esc(r['postal_code'])},{r['primary_region_seq']},{esc(r['area_type'])},{r['source_membership_count']},{esc('not_required' if r['source_membership_count']==1 else 'resolved')},{esc(r['source_region_seqs'])}::jsonb)" for r in postal_rows)
    tech_values=",".join(f"({esc(r['employee_code'])},{r['assigned_region_seq']},{esc(policy_mode)})" for r in tech_rows)
    if overflow:
        overflow_values=",".join(f"({esc(r['postal_code'])},{r['primary_region_seq']},{r['alternate_region_seq']},true,4500,{esc(r['rationale'])},{esc(policy_version)})" for r in overflow)
        overflow_guard=f"IF EXISTS (((SELECT postal_code,primary_region_seq,alternate_region_seq,allow_overflow,penalty_cost,rationale,policy_version FROM common_region_plan_boundary_overflow WHERE subsidiary_name='LGEAI' AND strategic_city_name={esc(city)} AND plan_id={esc(plan)}) EXCEPT (VALUES {overflow_values})) UNION ALL ((VALUES {overflow_values}) EXCEPT (SELECT postal_code,primary_region_seq,alternate_region_seq,allow_overflow,penalty_cost,rationale,policy_version FROM common_region_plan_boundary_overflow WHERE subsidiary_name='LGEAI' AND strategic_city_name={esc(city)} AND plan_id={esc(plan)}))) THEN RAISE EXCEPTION 'immutable plan overflow payload drift'; END IF;"
    else:
        overflow_guard=f"IF EXISTS (SELECT 1 FROM common_region_plan_boundary_overflow WHERE subsidiary_name='LGEAI' AND strategic_city_name={esc(city)} AND plan_id={esc(plan)}) THEN RAISE EXCEPTION 'immutable plan overflow payload drift'; END IF;"
    sql += [f"DO $$ BEGIN IF EXISTS (((SELECT {region_columns} FROM common_region_plan_region WHERE subsidiary_name='LGEAI' AND strategic_city_name={esc(city)} AND plan_id={esc(plan)}) EXCEPT (VALUES {region_values})) UNION ALL ((VALUES {region_values}) EXCEPT (SELECT {region_columns} FROM common_region_plan_region WHERE subsidiary_name='LGEAI' AND strategic_city_name={esc(city)} AND plan_id={esc(plan)}))) THEN RAISE EXCEPTION 'immutable plan region payload drift'; END IF; IF EXISTS (((SELECT postal_code,region_seq,area_type,source_membership_count,resolution_status,source_region_seqs FROM common_region_plan_postal WHERE subsidiary_name='LGEAI' AND strategic_city_name={esc(city)} AND plan_id={esc(plan)}) EXCEPT (VALUES {postal_values})) UNION ALL ((VALUES {postal_values}) EXCEPT (SELECT postal_code,region_seq,area_type,source_membership_count,resolution_status,source_region_seqs FROM common_region_plan_postal WHERE subsidiary_name='LGEAI' AND strategic_city_name={esc(city)} AND plan_id={esc(plan)}))) THEN RAISE EXCEPTION 'immutable plan postal payload drift'; END IF; IF EXISTS (((SELECT employee_code,assigned_region_seq,policy_mode FROM common_region_plan_technician WHERE subsidiary_name='LGEAI' AND strategic_city_name={esc(city)} AND plan_id={esc(plan)}) EXCEPT (VALUES {tech_values})) UNION ALL ((VALUES {tech_values}) EXCEPT (SELECT employee_code,assigned_region_seq,policy_mode FROM common_region_plan_technician WHERE subsidiary_name='LGEAI' AND strategic_city_name={esc(city)} AND plan_id={esc(plan)}))) THEN RAISE EXCEPTION 'immutable plan technician payload drift'; END IF; {overflow_guard} END $$;"]
    sql += [f"DO $$ BEGIN IF NOT EXISTS (SELECT 1 FROM common_region_plan WHERE subsidiary_name='LGEAI' AND strategic_city_name={esc(city)} AND plan_id={esc(plan)} AND plan_status='candidate') THEN RAISE EXCEPTION 'candidate plan lifecycle drift; review and activation must use the controlled repository workflow'; END IF; END $$;"]
    sql += ["COMMIT;"]
    (out/"import_vrp_db_dev.sql").write_text("\n".join(sql)+"\n",encoding="utf-8")
    return manifest

def main():
    ap=argparse.ArgumentParser(); ap.add_argument("--workbook",type=Path,required=True); ap.add_argument("--output-dir",type=Path,required=True); ap.add_argument("--target-city",required=True); ap.add_argument("--plan-id",required=True); ap.add_argument("--territory",action="append",required=True); ap.add_argument("--source-technician-city",required=True); ap.add_argument("--policy-version",choices=tuple(POLICY_MODES),required=True); ap.add_argument("--overflow-mode",choices=("explicit","disabled"),default="explicit"); args=ap.parse_args()
    build(args.workbook,args.output_dir,city=args.target_city,plan=args.plan_id,territory_order=args.territory,source_technician_city=args.source_technician_city,policy_version=args.policy_version,overflow_enabled=args.overflow_mode=="explicit")
if __name__ == "__main__": main()

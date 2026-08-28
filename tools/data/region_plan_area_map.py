"""Area Map -> common Region Plan v2 artifacts.

This module is intentionally UI and database independent.  Area Map supplies
the two operator-facing CSV files, this module validates and normalizes them
into the same ``Area``/``Technician`` workbook consumed by Region Plan v2, and
stores an auditable local candidate directory.  Database writes remain owned
by the Region Plan API and its review/activation lifecycle.
"""
from __future__ import annotations

import hashlib
import io
import json
import re
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Mapping

import pandas as pd
from openpyxl import Workbook, load_workbook

from tools.data.region_plan_workflow_v2 import canonicalize_workbook


POLICY_MODES = {
    "home_distance_only": "home_distance_only",
    "preferred_region_soft": "preferred_region_soft",
    "explicit_workbook_membership/v1": "assigned_region_boundary_spillover",
    "own_region_with_approved_boundary_overflow/v2": "assigned_region_boundary_spillover",
    "active_roster_type_hard_region_soft/v1": "active_roster_type_hard_region_soft",
    "active_roster_area_type_fallback_region_soft/v1": "active_roster_area_type_fallback_region_soft",
}
POLICY_VERSIONS = tuple(POLICY_MODES)
SAFE_PART = re.compile(r"^[A-Za-z0-9][A-Za-z0-9_.&() -]{0,159}$")
SAFE_ID = re.compile(r"^[A-Za-z0-9][A-Za-z0-9_.-]{0,79}$")
SAFE_CITY = re.compile(r"^[A-Za-z0-9][A-Za-z0-9_ .,&()/-]{0,159}$")

REGION_SOURCE_COLUMNS = (
    "POSTAL_CODE", "STRATEGIC_CITY_NAME", "region_id", "region_seq",
    "AREA_NAME", "new_region_name", "area_type",
)
TECH_SOURCE_COLUMNS = ("Tech ID", "Tech Name", "Assignment")
AREA_WORKBOOK_COLUMNS = (
    "region_code", "region_name", "region_seq", "postal_code", "area_type",
    "required_center_type", "membership_rank", "is_primary",
    "overflow_allowed", "overflow_penalty_minutes", "overflow_reason",
)
TECH_WORKBOOK_COLUMNS = (
    "technician_id", "region_code", "active", "policy_mode",
    "effective_from", "effective_to",
)


class AreaMapRegionPlanError(ValueError):
    """Stable validation error suitable for displaying in Streamlit."""

    def __init__(self, code: str, detail: str = "") -> None:
        self.code = code
        self.detail = detail
        super().__init__(f"{code}: {detail}" if detail else code)


@dataclass(frozen=True)
class AreaMapRegionPlanExport:
    workbook_bytes: bytes
    canonical_bytes: bytes
    rejects_bytes: bytes
    manifest: Mapping[str, Any]
    area_df: pd.DataFrame
    technician_df: pd.DataFrame
    source_region_bytes: bytes
    source_technician_bytes: bytes


def _text(value: object) -> str:
    if value is None:
        return ""
    value = str(value).strip()
    return "" if value.casefold() in {"nan", "none", "nat"} else value


def _sha(payload: bytes) -> str:
    return hashlib.sha256(payload).hexdigest()


def _safe_part(value: object, *, field: str, pattern: re.Pattern[str] = SAFE_PART) -> str:
    text = _text(value)
    if not text or not pattern.fullmatch(text):
        raise AreaMapRegionPlanError(f"{field.upper()}_INVALID")
    return text


def _city_folder_name(value: object) -> str:
    """Return one stable source-city directory name for all policy Plans."""
    text = _text(value)
    text = re.sub(r"\s+-\s+.*$", "", text).strip()
    text = re.sub(r"[^A-Za-z0-9]+", "_", text).strip("_")
    return _safe_part(text, field="source_city_folder", pattern=SAFE_ID)


def _read_table(file_name: str, payload: bytes) -> pd.DataFrame:
    if not payload:
        raise AreaMapRegionPlanError("SOURCE_FILE_EMPTY", Path(file_name).name)
    suffix = Path(file_name).suffix.casefold()
    try:
        if suffix == ".csv":
            # Excel copy/paste exports are often tab-delimited even when the
            # operator names the file with a .csv extension.  Detect that
            # common form while retaining normal comma CSV support.
            encoding = "utf-16" if payload.startswith((b"\xff\xfe", b"\xfe\xff")) else "utf-8-sig"
            sample = payload[:8192].decode(encoding, errors="replace")
            first_line = sample.splitlines()[0] if sample.splitlines() else ""
            delimiter_counts = {
                "\t": first_line.count("\t"),
                ",": first_line.count(","),
                ";": first_line.count(";"),
            }
            delimiter = max(delimiter_counts, key=delimiter_counts.get)
            return pd.read_csv(
                io.BytesIO(payload),
                sep=delimiter,
                encoding=encoding,
                dtype=str,
                keep_default_na=False,
            )
        if suffix in {".xlsx", ".xlsm"}:
            return pd.read_excel(io.BytesIO(payload), dtype=str, keep_default_na=False)
    except Exception as exc:  # pragma: no cover - provider/parser detail
        raise AreaMapRegionPlanError("SOURCE_FILE_INVALID", Path(file_name).name) from exc
    raise AreaMapRegionPlanError("SOURCE_EXTENSION_INVALID", Path(file_name).name)


def _header_map(frame: pd.DataFrame) -> dict[str, str]:
    result: dict[str, str] = {}
    for column in frame.columns:
        key = re.sub(r"[ _-]+", " ", _text(column)).casefold()
        result[key] = str(column)
    return result


def _column(frame: pd.DataFrame, *names: str) -> pd.Series:
    headers = _header_map(frame)
    for name in names:
        original = headers.get(re.sub(r"[ _-]+", " ", name).casefold())
        if original is not None:
            return frame[original].map(_text)
    return pd.Series([""] * len(frame), index=frame.index, dtype=str)


def _postal(value: object) -> str:
    value = _text(value).removesuffix(".0")
    if value.isdigit() and len(value) <= 5:
        return value.zfill(5)
    return value


def _write_workbook(area_df: pd.DataFrame, technician_df: pd.DataFrame) -> bytes:
    workbook = Workbook()
    area = workbook.active
    area.title = "Area"
    area.append(list(AREA_WORKBOOK_COLUMNS))
    for row in area_df.loc[:, list(AREA_WORKBOOK_COLUMNS)].itertuples(index=False, name=None):
        area.append(list(row))
    technician = workbook.create_sheet("Technician")
    technician.append(list(TECH_WORKBOOK_COLUMNS))
    for row in technician_df.loc[:, list(TECH_WORKBOOK_COLUMNS)].itertuples(index=False, name=None):
        technician.append(list(row))
    stream = io.BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def _validate_workbook(payload: bytes, metadata: Mapping[str, Any]) -> dict[str, Any]:
    # Loading once here catches malformed ZIPs before the canonical validator
    # and ensures the generated workbook is actually readable by Excel.
    try:
        workbook = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
        if set(workbook.sheetnames) != {"Area", "Technician"}:
            raise AreaMapRegionPlanError("GENERATED_WORKBOOK_SHEETS_INVALID")
    except AreaMapRegionPlanError:
        raise
    except Exception as exc:  # pragma: no cover - openpyxl parser detail
        raise AreaMapRegionPlanError("GENERATED_WORKBOOK_INVALID") from exc
    result = canonicalize_workbook(payload, metadata)
    manifest = result["manifest"]
    if manifest.get("status") != "candidate":
        errors = ",".join(map(str, manifest.get("plan_errors") or ()))
        raise AreaMapRegionPlanError("REGION_PLAN_VALIDATION_FAILED", errors)
    unexpected_rejects = [
        item for item in result.get("rejects") or ()
        if item.get("error_code") != "TECHNICIAN_ASSIGNMENT_BLANK"
    ]
    if unexpected_rejects:
        raise AreaMapRegionPlanError("REGION_PLAN_REJECTED_ROWS")
    return result


def build_area_map_region_plan(
    region_file_name: str,
    region_bytes: bytes,
    technician_file_name: str,
    technician_bytes: bytes,
    *,
    subsidiary_id: str,
    source_city_id: str,
    target_city_id: str,
    city_name: str = "",
    policy_version: str = "",
    plan_display_name: str = "",
    overflow_penalty_minutes: int = 4500,
) -> AreaMapRegionPlanExport:
    """Normalize Area Map source files into a common Region Plan workbook."""
    subsidiary_id = _safe_part(subsidiary_id, field="subsidiary_id", pattern=SAFE_ID)
    city_name = _safe_part(city_name or source_city_id or target_city_id, field="city_name", pattern=SAFE_CITY)
    source_city_id = city_name
    target_city_id = city_name
    try:
        penalty = int(overflow_penalty_minutes)
    except (TypeError, ValueError) as exc:
        raise AreaMapRegionPlanError("OVERFLOW_PENALTY_INVALID") from exc
    if penalty <= 0:
        raise AreaMapRegionPlanError("OVERFLOW_PENALTY_INVALID")

    region_source = _read_table(region_file_name, region_bytes)
    tech_source = _read_table(technician_file_name, technician_bytes)
    region_values = pd.DataFrame({
        "POSTAL_CODE": _column(region_source, "POSTAL_CODE", "postal_code", "ZIPCode"),
        "STRATEGIC_CITY_NAME": _column(region_source, "STRATEGIC_CITY_NAME", "strategic_city_name"),
        "region_id": _column(region_source, "region_id", "region code", "region_code"),
        "region_seq": _column(region_source, "region_seq", "region seq"),
        "AREA_NAME": _column(region_source, "AREA_NAME", "area name", "Territory"),
        "new_region_name": _column(region_source, "new_region_name", "region_name", "new region name"),
        "area_type": _column(region_source, "area_type", "Area Type"),
    })
    region_values = region_values[
        region_values.apply(lambda row: any(_text(value) for value in row), axis=1)
    ].reset_index(drop=True)
    if region_values.empty:
        raise AreaMapRegionPlanError("REGION_SOURCE_EMPTY")

    region_values["POSTAL_CODE"] = region_values["POSTAL_CODE"].map(_postal)
    region_values["area_type"] = region_values["area_type"].str.upper()
    region_values["new_region_name"] = region_values["new_region_name"].where(
        region_values["new_region_name"].ne(""), region_values["AREA_NAME"]
    )
    # New source files identify the roster city, while older Area Map exports
    # used the policy-city ID in this column.  Accept either form, but reject
    # any third/mixed city value instead of silently cross-wiring a plan.
    accepted_city_ids = {source_city_id, target_city_id}
    if (
        region_values["STRATEGIC_CITY_NAME"].eq("").any()
        or not region_values["STRATEGIC_CITY_NAME"].isin(accepted_city_ids).all()
    ):
        raise AreaMapRegionPlanError("REGION_CITY_MISMATCH")
    if (~region_values["POSTAL_CODE"].str.fullmatch(r"\d{5}")).any():
        raise AreaMapRegionPlanError("POSTAL_CODE_INVALID")
    required_region_fields = ["region_id", "region_seq", "AREA_NAME", "new_region_name", "area_type"]
    missing_fields = [
        field for field in required_region_fields
        if region_values[field].eq("").any()
    ]
    if missing_fields:
        raise AreaMapRegionPlanError(
            "REGION_FIELD_MISSING",
            "missing=" + ",".join(missing_fields),
        )
    try:
        region_values["region_seq"] = region_values["region_seq"].map(lambda value: int(float(value)))
    except (TypeError, ValueError) as exc:
        raise AreaMapRegionPlanError("REGION_SEQ_INVALID") from exc
    if (region_values["region_seq"] <= 0).any():
        raise AreaMapRegionPlanError("REGION_SEQ_INVALID")
    if (~region_values["area_type"].isin({"DMS", "DMS2"})).any():
        raise AreaMapRegionPlanError("AREA_TYPE_INVALID")

    normalization_warnings: list[str] = []
    original_sequences = sorted(int(value) for value in region_values["region_seq"].unique())
    expected_sequences = list(range(1, len(original_sequences) + 1))
    sequence_aliases: dict[str, str] = {}
    if original_sequences != expected_sequences:
        sequence_map = {
            old_sequence: new_sequence
            for new_sequence, old_sequence in enumerate(original_sequences, start=1)
        }
        sequence_aliases = {
            str(old_sequence): str(new_sequence)
            for old_sequence, new_sequence in sequence_map.items()
        }
        region_values["region_seq"] = region_values["region_seq"].map(sequence_map)
        normalization_warnings.append(
            "region_seq normalized from "
            f"{original_sequences} to {expected_sequences}"
        )
    region_registry = region_values[["region_id", "region_seq", "AREA_NAME", "new_region_name", "area_type"]].drop_duplicates()
    id_seq_conflicts = set(
        region_registry.groupby("region_id")["region_seq"].nunique().loc[lambda values: values.gt(1)].index
    )
    seq_id_conflicts = set(
        region_registry.groupby("region_seq")["region_id"].nunique().loc[lambda values: values.gt(1)].index
    )
    if id_seq_conflicts or seq_id_conflicts:
        # Postal rows are the membership source of truth and region_seq is the
        # stable region key.  Operators commonly move a postal code to a new
        # Zone but leave the old region_id in that row.  Preserve consistent
        # IDs, while assigning deterministic IDs to only the conflicted
        # sequences.  Technician assignments by Zone/name/sequence continue
        # to resolve through the lookup below; an ambiguous old ID is rejected
        # instead of silently assigning a technician to the wrong region.
        for region_seq in sorted(seq_id_conflicts | {
            int(value)
            for region_id in id_seq_conflicts
            for value in region_registry.loc[region_registry["region_id"].eq(region_id), "region_seq"].tolist()
        }):
            generated_id = f"{target_city_id}_r{int(region_seq):02d}"
            region_values.loc[region_values["region_seq"].eq(region_seq), "region_id"] = generated_id
            normalization_warnings.append(
                f"region_seq {int(region_seq)} received normalized region_id {generated_id}"
            )
        region_registry = region_values[["region_id", "region_seq", "AREA_NAME", "new_region_name", "area_type"]].drop_duplicates()
    if set(region_values["region_seq"]) != set(range(1, int(region_values["region_seq"].max()) + 1)):
        raise AreaMapRegionPlanError("REGION_SEQ_NOT_CONTIGUOUS")

    region_values["membership_rank"] = region_values.groupby("POSTAL_CODE", sort=False).cumcount() + 1
    region_values["is_primary"] = region_values["membership_rank"].eq(1)
    duplicate_mask = region_values["POSTAL_CODE"].duplicated(keep=False)
    region_values["overflow_allowed"] = duplicate_mask & ~region_values["is_primary"]
    region_values["overflow_penalty_minutes"] = region_values["overflow_allowed"].map(lambda value: penalty if value else "")
    region_values["overflow_reason"] = region_values["overflow_allowed"].map(
        lambda value: "explicit duplicate postal membership from Area Map source" if value else ""
    )
    area_df = region_values.rename(columns={
        "POSTAL_CODE": "postal_code", "region_id": "region_code", "new_region_name": "region_name",
        "area_type": "area_type",
    })[
        ["region_code", "region_name", "region_seq", "postal_code", "area_type",
         "membership_rank", "is_primary", "overflow_allowed", "overflow_penalty_minutes", "overflow_reason"]
    ].copy()
    area_df["required_center_type"] = area_df["area_type"]
    area_df = area_df.loc[:, list(AREA_WORKBOOK_COLUMNS)]

    tech_values = pd.DataFrame({
        "technician_id": _column(tech_source, "Tech ID", "technician_id", "employee_code"),
        "technician_name": _column(tech_source, "Tech Name", "technician_name", "employee_name"),
        "assignment": _column(tech_source, "Assignment", "region_code", "region_id"),
    })
    tech_values = tech_values[
        tech_values.apply(lambda row: any(_text(value) for value in row), axis=1)
    ].reset_index(drop=True)
    if tech_values.empty:
        raise AreaMapRegionPlanError("TECHNICIAN_SOURCE_EMPTY")
    tech_values["technician_id"] = tech_values["technician_id"].str.upper()
    if tech_values["technician_id"].eq("").any():
        raise AreaMapRegionPlanError("TECHNICIAN_FIELD_MISSING")
    if tech_values["technician_id"].duplicated().any():
        raise AreaMapRegionPlanError("TECHNICIAN_ID_DUPLICATE")
    assignment_lookup: dict[str, str] = {}
    for row in region_registry.itertuples(index=False):
        code, seq, area_name, display_name, _area_type = map(_text, row)
        for key in (code, area_name, display_name, str(seq)):
            if key:
                assignment_lookup[key.casefold()] = code
    for original_sequence, normalized_sequence in sequence_aliases.items():
        code = assignment_lookup.get(normalized_sequence.casefold())
        if code:
            assignment_lookup[original_sequence.casefold()] = code
    tech_values["region_code"] = tech_values["assignment"].map(
        lambda value: assignment_lookup.get(_text(value).casefold(), "")
    )
    unknown_assignment = tech_values["assignment"].ne("") & tech_values["region_code"].eq("")
    if unknown_assignment.any():
        raise AreaMapRegionPlanError("TECHNICIAN_ASSIGNMENT_UNKNOWN")
    technician_df = tech_values.assign(
        active="true", policy_mode="", effective_from="", effective_to=""
    )[["technician_id", "region_code", "active", "policy_mode", "effective_from", "effective_to"]]

    metadata = {
        "subsidiary_id": subsidiary_id,
        "city_name": city_name,
        "target_city_id": target_city_id,
        "source_city_id": source_city_id,
        "overlap_policy": "area_map_duplicate_postal_explicit_overflow",
        "activation_intent": "review_only",
        "plan_display_name": str(plan_display_name).strip() or target_city_id,
    }
    workbook_bytes = _write_workbook(area_df, technician_df)
    result = _validate_workbook(workbook_bytes, metadata)
    manifest = dict(result["manifest"])
    manifest.update({
        "source": {
            "region_file_name": Path(region_file_name).name,
            "technician_file_name": Path(technician_file_name).name,
            "region_sha256": _sha(region_bytes),
            "technician_sha256": _sha(technician_bytes),
        },
        "generated_workbook_sha256": _sha(workbook_bytes),
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "generator": "area-map-region-plan/v1",
        "privacy": "Technician names are source-only and excluded from canonical artifacts.",
        "normalization_warnings": normalization_warnings,
    })
    return AreaMapRegionPlanExport(
        workbook_bytes=workbook_bytes,
        canonical_bytes=result["artifacts"]["canonical.json"],
        rejects_bytes=result["artifacts"]["rejects.jsonl"],
        manifest=manifest,
        area_df=area_df,
        technician_df=technician_df,
        source_region_bytes=region_bytes,
        source_technician_bytes=technician_bytes,
    )


def save_area_map_region_plan(export: AreaMapRegionPlanExport, root: Path) -> Path:
    """Persist one deterministic local candidate and return its directory."""
    metadata = export.manifest.get("city_metadata") or {}
    subsidiary = _safe_part(metadata.get("subsidiary_id"), field="subsidiary_id", pattern=SAFE_ID)
    city = _city_folder_name(
        metadata.get("source_city_id") or metadata.get("target_city_id")
    )
    plan_id = _safe_part(export.manifest.get("plan_id"), field="plan_id", pattern=SAFE_ID)
    directory = (root / subsidiary / city / plan_id).resolve()
    root = root.resolve()
    if root != directory and root not in directory.parents:
        raise AreaMapRegionPlanError("REGION_PLAN_PATH_INVALID")
    (directory / "source").mkdir(parents=True, exist_ok=True)
    (directory / "normalized").mkdir(parents=True, exist_ok=True)
    (directory / "source" / str((export.manifest.get("source") or {}).get("region_file_name") or "region.csv")).write_bytes(export.source_region_bytes)
    (directory / "source" / str((export.manifest.get("source") or {}).get("technician_file_name") or "technician.csv")).write_bytes(export.source_technician_bytes)
    export.area_df.to_csv(directory / "normalized" / "area.csv", index=False, encoding="utf-8-sig")
    export.technician_df.to_csv(directory / "normalized" / "technician.csv", index=False, encoding="utf-8-sig")
    (directory / "region_plan.xlsx").write_bytes(export.workbook_bytes)
    manifest_bytes = json.dumps(export.manifest, ensure_ascii=False, indent=2, sort_keys=True).encode("utf-8")
    (directory / "manifest.json").write_bytes(manifest_bytes)
    (directory / "canonical.json").write_bytes(export.canonical_bytes)
    (directory / "rejects.jsonl").write_bytes(export.rejects_bytes)
    (directory / "checksums.json").write_text(json.dumps({
        "region_plan.xlsx": _sha(export.workbook_bytes),
        "manifest.json": _sha(manifest_bytes),
        "canonical.json": _sha(export.canonical_bytes),
        "rejects.jsonl": _sha(export.rejects_bytes),
        "normalized/area.csv": _sha((directory / "normalized" / "area.csv").read_bytes()),
        "normalized/technician.csv": _sha((directory / "normalized" / "technician.csv").read_bytes()),
    }, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    return directory


def list_saved_region_plan_workbooks(root: Path) -> list[dict[str, Any]]:
    """Return safe metadata for workbooks generated by Area Map."""
    root = root.resolve()
    if not root.is_dir():
        return []
    result: list[dict[str, Any]] = []
    for workbook_path in sorted(root.glob("*/*/*/region_plan.xlsx")):
        try:
            workbook_path = workbook_path.resolve()
            if root not in workbook_path.parents or not workbook_path.is_file():
                continue
            manifest_path = workbook_path.parent / "manifest.json"
            manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
            metadata = manifest.get("city_metadata") or {}
            display_name = str(manifest.get("plan_display_name") or "").strip()
            target_name = str(metadata.get("target_city_id", workbook_path.parent.parent.name)).strip()
            plan_id = str(manifest.get("plan_id", workbook_path.parent.name)).strip()
            label = f"{display_name} | {target_name} | {plan_id}" if display_name else f"{target_name} | {plan_id}"
            result.append({
                "label": label,
                "path": workbook_path,
                "manifest": manifest,
            })
        except (OSError, ValueError, TypeError, json.JSONDecodeError):
            continue
    return result


__all__ = [
    "AreaMapRegionPlanError", "AreaMapRegionPlanExport", "POLICY_MODES", "POLICY_VERSIONS",
    "build_area_map_region_plan", "list_saved_region_plan_workbooks", "save_area_map_region_plan",
]

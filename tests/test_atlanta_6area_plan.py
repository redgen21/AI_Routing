import csv
import hashlib
import io
import json
import unittest
import zipfile
from pathlib import Path

import openpyxl

from tools.data.atlanta_6area_plan import (
    AREA_SHEET_NAME,
    BOUNDARY_PENALTY_COST,
    BOUNDARY_POLICY_FILENAME,
    EXPECTED_AMBIGUOUS_POSTALS,
    EXPECTED_SOURCE_SHA256,
    FIXED_REGION_FILENAME,
    MANIFEST_FILENAME,
    MANIFEST_SCHEMA,
    PLAN_ID,
    POLICY_VERSION,
    TECHNICIAN_SHEET_NAME,
    TECHNICIAN_POLICY_FILENAME,
    Atlanta6AreaPlanError,
    build_atlanta_6area_bundle,
    build_atlanta_6area_template_bytes,
    derive_atlanta_6area_plan_identity,
    get_atlanta_6area_template_bytes,
    get_atlanta_6area_workbook_template,
    parse_atlanta_6area_workbook,
    preview_atlanta_6area_plan,
    validate_atlanta_6area_bundle,
)


PROJECT_ROOT = Path(__file__).resolve().parents[1]
SOURCE = PROJECT_ROOT / "260310" / "New ATL Buckets.xlsx"
EXISTING_FIXED_REGION = (
    PROJECT_ROOT
    / "data"
    / "north_america"
    / "reviewed"
    / "regions"
    / "fixed_region_postal_atlanta_6area_6_dms_v2.csv"
)


def _boundary_resolutions() -> dict[str, dict[str, object]]:
    return {
        postal: {"primary_region": "Zone 3", "allow_overflow": True}
        for postal in EXPECTED_AMBIGUOUS_POSTALS
    }


def _csv_rows(payload: bytes) -> list[dict[str, str]]:
    return list(csv.DictReader(io.StringIO(payload.decode("utf-8-sig"))))


def _mutated_source(sheet_name: str, mutator) -> bytes:
    workbook = openpyxl.load_workbook(SOURCE)
    try:
        mutator(workbook[sheet_name])
        stream = io.BytesIO()
        workbook.save(stream)
        return stream.getvalue()
    finally:
        workbook.close()


class Atlanta6AreaPlanTests(unittest.TestCase):
    def test_preview_accounts_for_split_source_without_exposing_technicians(self) -> None:
        preview = preview_atlanta_6area_plan(SOURCE)
        summary = preview.summary
        self.assertFalse(preview.promotable)
        self.assertEqual(preview.approval_status, "pending_boundary_resolutions")
        self.assertEqual(summary["membership_input_rows"], 301)
        self.assertEqual(summary["membership_accepted_rows"], 301)
        self.assertEqual(summary["membership_rejected_rows"], 0)
        self.assertEqual(summary["unique_postal_count"], 297)
        self.assertEqual(summary["ambiguous_postals"], list(EXPECTED_AMBIGUOUS_POSTALS))
        self.assertEqual(summary["technician_input_rows"], 14)
        self.assertEqual(summary["source_sha256"], EXPECTED_SOURCE_SHA256)
        self.assertEqual(
            summary["territory_membership_counts"],
            {
                "Zone 1": 73,
                "Zone 2": 46,
                "Zone 3": 25,
                "Zone 4": 49,
                "Zone 5": 65,
                "ATL Outer Area": 43,
            },
        )
        rendered = json.dumps(preview.as_dict())
        self.assertNotIn("AI105115", rendered)
        self.assertNotIn("Jason Patterson", rendered)

    def test_bundle_applies_zip_level_owners_and_only_enabled_overflow_policies(self) -> None:
        bundle = build_atlanta_6area_bundle(
            SOURCE, boundary_resolutions=_boundary_resolutions()
        )
        self.assertFalse(bundle.promotable)
        self.assertRegex(bundle.plan_id, r"^atlanta_6area_v2_[0-9a-f]{64}$")
        self.assertEqual(
            bundle.plan_id,
            "atlanta_6area_v2_51a40fc9ddf29b6b2e095050ecc24a72e874b9f93f383b33f55360f21f3202d2",
        )
        self.assertEqual(
            bundle.approval_status, "resolved_for_development_verification"
        )
        self.assertEqual(
            set(bundle.artifacts),
            {
                FIXED_REGION_FILENAME,
                BOUNDARY_POLICY_FILENAME,
                TECHNICIAN_POLICY_FILENAME,
                MANIFEST_FILENAME,
            },
        )

        fixed = _csv_rows(bundle.artifacts[FIXED_REGION_FILENAME])
        self.assertEqual(len(fixed), 297)
        self.assertEqual(len({row["POSTAL_CODE"] for row in fixed}), 297)
        self.assertEqual({row["area_type"] for row in fixed}, {"DMS"})
        boundary_fixed = [row for row in fixed if row["POSTAL_CODE"] in EXPECTED_AMBIGUOUS_POSTALS]
        self.assertEqual(
            {row["POSTAL_CODE"]: row["region_seq"] for row in boundary_fixed},
            {postal: "3" for postal in EXPECTED_AMBIGUOUS_POSTALS},
        )
        self.assertEqual({row["STRATEGIC_CITY_NAME"] for row in fixed}, {"Atlanta_6area"})
        self.assertEqual(
            sorted({(row["region_id"], row["region_seq"]) for row in fixed}),
            [
                ("atlanta_6area_r01", "1"),
                ("atlanta_6area_r02", "2"),
                ("atlanta_6area_r03", "3"),
                ("atlanta_6area_r04", "4"),
                ("atlanta_6area_r05", "5"),
                ("atlanta_6area_r06", "6"),
            ],
        )

        boundary = _csv_rows(bundle.artifacts[BOUNDARY_POLICY_FILENAME])
        self.assertEqual({row["plan_id"] for row in boundary}, {bundle.plan_id})
        self.assertEqual([row["POSTAL_CODE"] for row in boundary], list(EXPECTED_AMBIGUOUS_POSTALS))
        self.assertEqual(
            {row["POSTAL_CODE"]: row["primary_region_seq"] for row in boundary},
            {postal: "3" for postal in EXPECTED_AMBIGUOUS_POSTALS},
        )
        self.assertEqual(
            {row["POSTAL_CODE"]: row["alternate_region_seq"] for row in boundary},
            {postal: "2" for postal in EXPECTED_AMBIGUOUS_POSTALS},
        )
        self.assertEqual({row["penalty_cost"] for row in boundary}, {str(BOUNDARY_PENALTY_COST)})

        technicians = _csv_rows(bundle.artifacts[TECHNICIAN_POLICY_FILENAME])
        self.assertEqual({row["plan_id"] for row in technicians}, {bundle.plan_id})
        self.assertEqual(len(technicians), 14)
        self.assertEqual(len({row["SVC_ENGINEER_CODE"] for row in technicians}), 14)
        self.assertNotIn("SVC_ENGINEER_NAME", technicians[0])
        self.assertEqual(
            {row["policy_mode"] for row in technicians},
            {"assigned_region_boundary_spillover"},
        )
        outer_area_technician = next(
            row for row in technicians if row["SVC_ENGINEER_CODE"] == "AI105115"
        )
        self.assertEqual(outer_area_technician["assigned_region_seq"], "6")

        manifest = json.loads(bundle.artifacts[MANIFEST_FILENAME].decode("utf-8"))
        self.assertEqual(manifest["plan_id"], bundle.plan_id)
        self.assertEqual(manifest["schema_version"], MANIFEST_SCHEMA)
        self.assertEqual(manifest["policy_version"], POLICY_VERSION)
        self.assertEqual(bundle.plan_id, f"atlanta_6area_v2_{manifest['resolution_digest']}")
        self.assertEqual(
            manifest["approval_status"], "resolved_for_development_verification"
        )
        self.assertFalse(manifest["promotable"])
        self.assertTrue(manifest["verification_only"])
        self.assertEqual(manifest["lifecycle_stage"], "resolved_candidate")
        self.assertEqual(manifest["privacy_classification"], "internal_pii_redacted")
        self.assertTrue(manifest["technician_names_redacted"])
        self.assertEqual(
            manifest["source"],
            {
                "file_name": "New ATL Buckets.xlsx",
                "sha256": EXPECTED_SOURCE_SHA256,
                "sheets": [
                    {
                        "role": "membership",
                        "sheet_name": AREA_SHEET_NAME,
                        "headers": ["ZIPCode", "Territory"],
                        "input_rows": 301,
                    },
                    {
                        "role": "technician",
                        "sheet_name": TECHNICIAN_SHEET_NAME,
                        "headers": ["Tech ID", "Tech Name", "Assignment"],
                        "input_rows": 14,
                    },
                ],
            },
        )
        self.assertEqual(
            manifest["source_technician_master_context"],
            {"subsidiary_name": "LGEAI", "strategic_city_name": "Atlanta, GA"},
        )
        self.assertEqual(
            manifest["target_technician_master_context"],
            {"subsidiary_name": "LGEAI", "strategic_city_name": "Atlanta_6area"},
        )
        self.assertEqual(manifest["row_accounting"]["membership"]["input"], 301)
        self.assertEqual(
            manifest["boundary_resolutions"]["30028"],
            {
                "allow_overflow": True,
                "alternate_region": "Zone 2",
                "alternate_region_seq": 2,
                "penalty_cost": BOUNDARY_PENALTY_COST,
                "primary_region": "Zone 3",
                "primary_region_seq": 3,
                "rationale": "",
            },
        )
        self.assertEqual(
            manifest["boundary_resolutions"]["30040"]["penalty_cost"],
            BOUNDARY_PENALTY_COST,
        )
        for name in (FIXED_REGION_FILENAME, BOUNDARY_POLICY_FILENAME, TECHNICIAN_POLICY_FILENAME):
            self.assertEqual(
                manifest["artifacts"][name]["sha256"],
                hashlib.sha256(bundle.artifacts[name]).hexdigest(),
            )

    def test_zone_3_overlap_decisions_reproduce_existing_fixed_region_projection(self) -> None:
        bundle = build_atlanta_6area_bundle(
            SOURCE, boundary_resolutions=_boundary_resolutions()
        )
        with EXISTING_FIXED_REGION.open(encoding="utf-8-sig", newline="") as stream:
            expected_rows = list(csv.DictReader(stream))
        self.assertEqual(_csv_rows(bundle.artifacts[FIXED_REGION_FILENAME]), expected_rows)

    def test_bundle_redacts_source_technician_names_from_all_artifacts(self) -> None:
        synthetic_name = "Synthetic Private Technician"
        source = _mutated_source(
            TECHNICIAN_SHEET_NAME,
            lambda sheet: setattr(sheet["B2"], "value", synthetic_name),
        )
        bundle = build_atlanta_6area_bundle(
            source, boundary_resolutions=_boundary_resolutions()
        )
        technician_csv = bundle.artifacts[TECHNICIAN_POLICY_FILENAME]
        self.assertNotIn(b"SVC_ENGINEER_NAME", technician_csv)
        self.assertNotIn(synthetic_name.encode("utf-8"), technician_csv)
        self.assertNotIn(synthetic_name.encode("utf-8"), bundle.bundle_bytes)
        self.assertTrue(bundle.manifest["technician_names_redacted"])
        self.assertEqual(bundle.manifest["privacy_classification"], "internal_pii_redacted")
        self.assertNotEqual(bundle.manifest["source"]["sha256"], EXPECTED_SOURCE_SHA256)
        self.assertEqual(
            validate_atlanta_6area_bundle(bundle.bundle_bytes).plan_id, bundle.plan_id
        )

    def test_template_download_is_deterministic_blank_and_exactly_split(self) -> None:
        payload = build_atlanta_6area_template_bytes()
        self.assertEqual(payload, build_atlanta_6area_template_bytes())
        self.assertEqual(payload, get_atlanta_6area_template_bytes())
        self.assertEqual(payload, get_atlanta_6area_workbook_template())
        workbook = openpyxl.load_workbook(io.BytesIO(payload))
        try:
            self.assertEqual(workbook.sheetnames, [AREA_SHEET_NAME, TECHNICIAN_SHEET_NAME])
            self.assertEqual(
                tuple(cell.value for cell in workbook[AREA_SHEET_NAME][1]),
                ("ZIPCode", "Territory"),
            )
            self.assertEqual(
                tuple(cell.value for cell in workbook[TECHNICIAN_SHEET_NAME][1]),
                ("Tech ID", "Tech Name", "Assignment"),
            )
            self.assertTrue(workbook[AREA_SHEET_NAME].data_validations.dataValidation)
            self.assertTrue(workbook[TECHNICIAN_SHEET_NAME].data_validations.dataValidation)
        finally:
            workbook.close()
        with self.assertRaisesRegex(Atlanta6AreaPlanError, "MEMBERSHIP_ROW_COUNT_INVALID"):
            parse_atlanta_6area_workbook(payload)

    def test_bundle_is_deterministic_for_bytes_and_path(self) -> None:
        resolutions = _boundary_resolutions()
        first = build_atlanta_6area_bundle(SOURCE, boundary_resolutions=resolutions)
        second = build_atlanta_6area_bundle(
            SOURCE.read_bytes(), boundary_resolutions=resolutions
        )
        self.assertEqual(first.bundle_bytes, second.bundle_bytes)
        self.assertEqual(dict(first.artifacts), dict(second.artifacts))
        with zipfile.ZipFile(io.BytesIO(first.bundle_bytes)) as archive:
            self.assertEqual(sorted(archive.namelist()), sorted(first.artifacts))
            for name, payload in first.artifacts.items():
                self.assertEqual(archive.read(name), payload)

    def test_bundle_validator_rechecks_canonical_csv_manifest_hashes_and_accounting(self) -> None:
        bundle = build_atlanta_6area_bundle(
            SOURCE, boundary_resolutions=_boundary_resolutions()
        )
        validated = validate_atlanta_6area_bundle(bundle.bundle_bytes)
        self.assertEqual(validated.plan_id, bundle.plan_id)
        self.assertEqual(validated.bundle_bytes, bundle.bundle_bytes)
        self.assertEqual(validated.summary["canonical_region_count"], 6)
        self.assertEqual(validated.summary["canonical_postal_count"], 297)
        self.assertEqual(validated.summary["canonical_technician_count"], 14)
        self.assertEqual(validated.summary["canonical_boundary_policy_count"], 4)

        altered = dict(bundle.artifacts)
        altered[FIXED_REGION_FILENAME] = altered[FIXED_REGION_FILENAME].replace(
            b"30028", b"99999", 1
        )
        stream = io.BytesIO()
        with zipfile.ZipFile(stream, "w", compression=zipfile.ZIP_DEFLATED) as archive:
            for name in sorted(altered):
                archive.writestr(name, altered[name])
        with self.assertRaisesRegex(Atlanta6AreaPlanError, "BUNDLE_ARTIFACT_HASH_INVALID"):
            validate_atlanta_6area_bundle(stream.getvalue())

    def test_validator_rejects_blank_fixed_region_area_type(self) -> None:
        bundle = build_atlanta_6area_bundle(
            SOURCE, boundary_resolutions=_boundary_resolutions()
        )
        altered = dict(bundle.artifacts)
        rows = _csv_rows(altered[FIXED_REGION_FILENAME])
        rows[0]["area_type"] = ""
        stream = io.StringIO(newline="")
        writer = csv.DictWriter(
            stream, fieldnames=tuple(rows[0]), lineterminator="\n"
        )
        writer.writeheader()
        writer.writerows(rows)
        altered[FIXED_REGION_FILENAME] = stream.getvalue().encode("utf-8-sig")
        manifest = json.loads(altered[MANIFEST_FILENAME].decode("utf-8"))
        manifest["artifacts"][FIXED_REGION_FILENAME] = {
            "sha256": hashlib.sha256(altered[FIXED_REGION_FILENAME]).hexdigest(),
            "size_bytes": len(altered[FIXED_REGION_FILENAME]),
            "row_count": len(rows),
        }
        altered[MANIFEST_FILENAME] = (
            json.dumps(manifest, ensure_ascii=False, indent=2, sort_keys=True) + "\n"
        ).encode("utf-8")
        archive_stream = io.BytesIO()
        with zipfile.ZipFile(
            archive_stream, "w", compression=zipfile.ZIP_DEFLATED
        ) as archive:
            for name in sorted(altered):
                archive.writestr(name, altered[name])
        with self.assertRaisesRegex(
            Atlanta6AreaPlanError, "BUNDLE_FIXED_REGION_VALUE_INVALID"
        ):
            validate_atlanta_6area_bundle(archive_stream.getvalue())

    def test_plan_identity_binds_source_sorted_decisions_policy_and_schema(self) -> None:
        parsed = parse_atlanta_6area_workbook(SOURCE)
        decisions = _boundary_resolutions()
        digest, plan_id = derive_atlanta_6area_plan_identity(parsed.source_sha256, decisions)
        reversed_digest, reversed_id = derive_atlanta_6area_plan_identity(
            parsed.source_sha256, dict(reversed(list(decisions.items())))
        )
        self.assertEqual((reversed_digest, reversed_id), (digest, plan_id))
        self.assertEqual(plan_id, f"atlanta_6area_v2_{digest}")
        self.assertEqual(
            plan_id,
            "atlanta_6area_v2_51a40fc9ddf29b6b2e095050ecc24a72e874b9f93f383b33f55360f21f3202d2",
        )

        changed = _boundary_resolutions()
        changed["30040"]["allow_overflow"] = False
        self.assertNotEqual(
            derive_atlanta_6area_plan_identity(parsed.source_sha256, changed)[0], digest
        )
        self.assertNotEqual(
            derive_atlanta_6area_plan_identity("0" * 64, decisions)[0], digest
        )

        bundle = build_atlanta_6area_bundle(SOURCE, boundary_resolutions=decisions)
        manifest = bundle.manifest
        identity = {
            "source_sha256": parsed.source_sha256,
            "boundary_resolutions": [
                {"postal_code": postal, **dict(manifest["boundary_resolutions"][postal])}
                for postal in sorted(manifest["boundary_resolutions"])
            ],
            "policy_version": POLICY_VERSION,
            "schema_version": MANIFEST_SCHEMA,
            "source_technician_master_context": {
                "subsidiary_name": "LGEAI", "strategic_city_name": "Atlanta, GA"
            },
            "target_technician_master_context": {
                "subsidiary_name": "LGEAI", "strategic_city_name": "Atlanta_6area"
            },
            "technician_names_redacted": True,
        }
        expected = hashlib.sha256(
            json.dumps(identity, ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode("utf-8")
        ).hexdigest()
        self.assertEqual(digest, expected)

    def test_validator_rejects_technician_policy_containing_a_name_column(self) -> None:
        bundle = build_atlanta_6area_bundle(
            SOURCE, boundary_resolutions=_boundary_resolutions()
        )
        altered = dict(bundle.artifacts)
        rows = _csv_rows(altered[TECHNICIAN_POLICY_FILENAME])
        columns = tuple(rows[0]) + ("SVC_ENGINEER_NAME",)
        for row in rows:
            row["SVC_ENGINEER_NAME"] = "Synthetic Private Technician"
        stream = io.StringIO(newline="")
        writer = csv.DictWriter(stream, fieldnames=columns, lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)
        altered[TECHNICIAN_POLICY_FILENAME] = stream.getvalue().encode("utf-8-sig")
        manifest = json.loads(altered[MANIFEST_FILENAME].decode("utf-8"))
        manifest["artifacts"][TECHNICIAN_POLICY_FILENAME] = {
            "sha256": hashlib.sha256(altered[TECHNICIAN_POLICY_FILENAME]).hexdigest(),
            "size_bytes": len(altered[TECHNICIAN_POLICY_FILENAME]),
            "row_count": len(rows),
        }
        altered[MANIFEST_FILENAME] = (
            json.dumps(manifest, ensure_ascii=False, indent=2, sort_keys=True) + "\n"
        ).encode("utf-8")
        archive_stream = io.BytesIO()
        with zipfile.ZipFile(archive_stream, "w", compression=zipfile.ZIP_DEFLATED) as archive:
            for name in sorted(altered):
                archive.writestr(name, altered[name])
        with self.assertRaisesRegex(Atlanta6AreaPlanError, "BUNDLE_ARTIFACT_SCHEMA_INVALID"):
            validate_atlanta_6area_bundle(archive_stream.getvalue())

    def test_bundle_requires_exact_valid_zip_level_resolutions(self) -> None:
        cases: list[tuple[str, object]] = []
        missing = _boundary_resolutions()
        missing.pop("30028")
        cases.append(("BOUNDARY_RESOLUTION_MISSING", missing))
        unknown = _boundary_resolutions()
        unknown["99999"] = {"primary_region": "Zone 2", "allow_overflow": False}
        cases.append(("BOUNDARY_RESOLUTION_UNKNOWN", unknown))
        invalid_primary = _boundary_resolutions()
        invalid_primary["30028"]["primary_region"] = "Zone 1"
        cases.append(("BOUNDARY_PRIMARY_REGION_INVALID", invalid_primary))
        invalid_overflow = _boundary_resolutions()
        invalid_overflow["30028"]["allow_overflow"] = "yes"
        cases.append(("BOUNDARY_ALLOW_OVERFLOW_INVALID", invalid_overflow))
        cases.append(("BOUNDARY_RESOLUTIONS_REQUIRED", None))
        for expected, value in cases:
            with self.subTest(expected=expected):
                with self.assertRaisesRegex(Atlanta6AreaPlanError, expected):
                    build_atlanta_6area_bundle(
                        SOURCE, boundary_resolutions=value  # type: ignore[arg-type]
                    )

    def test_rejects_unknown_duplicate_and_missing_technician_ids(self) -> None:
        cases = (
            (
                "TECH_ID_UNKNOWN",
                lambda sheet: setattr(sheet["A2"], "value", "ZZ999999"),
            ),
            (
                "TECH_ID_DUPLICATE",
                lambda sheet: setattr(sheet["A3"], "value", sheet["A2"].value),
            ),
            (
                "TECH_ID_MISSING",
                lambda sheet: setattr(sheet["A2"], "value", None),
            ),
        )
        for expected, mutator in cases:
            with self.subTest(expected=expected):
                with self.assertRaisesRegex(Atlanta6AreaPlanError, expected):
                    parse_atlanta_6area_workbook(
                        _mutated_source(TECHNICIAN_SHEET_NAME, mutator)
                    )

    def test_required_source_fields_fail_without_silent_defaults(self) -> None:
        cases = (
            (
                "POSTAL_MEMBERSHIP_INCOMPLETE",
                "row=2,field=ZIPCode",
                AREA_SHEET_NAME,
                lambda sheet: setattr(sheet["A2"], "value", None),
            ),
            (
                "POSTAL_MEMBERSHIP_INCOMPLETE",
                "row=2,field=Territory",
                AREA_SHEET_NAME,
                lambda sheet: setattr(sheet["B2"], "value", None),
            ),
            (
                "TECH_ID_MISSING",
                "row=2,field=Tech ID",
                TECHNICIAN_SHEET_NAME,
                lambda sheet: setattr(sheet["A2"], "value", None),
            ),
            (
                "TECH_ASSIGNMENT_INVALID",
                "row=2,field=Assignment",
                TECHNICIAN_SHEET_NAME,
                lambda sheet: setattr(sheet["C2"], "value", None),
            ),
        )
        for code, detail, sheet_name, mutator in cases:
            with self.subTest(code=code, detail=detail):
                with self.assertRaises(Atlanta6AreaPlanError) as raised:
                    parse_atlanta_6area_workbook(_mutated_source(sheet_name, mutator))
                self.assertEqual(raised.exception.code, code)
                self.assertEqual(raised.exception.detail, detail)

    def test_allows_blank_rows_but_rejects_populated_trailing_cells(self) -> None:
        blank_rows = _mutated_source(
            AREA_SHEET_NAME,
            lambda sheet: sheet.cell(row=500, column=1, value=None),
        )
        self.assertEqual(
            parse_atlanta_6area_workbook(blank_rows).source_sha256,
            hashlib.sha256(blank_rows).hexdigest(),
        )
        trailing = _mutated_source(
            AREA_SHEET_NAME,
            lambda sheet: sheet.cell(row=2, column=3, value="unexpected"),
        )
        with self.assertRaises(Atlanta6AreaPlanError) as raised:
            parse_atlanta_6area_workbook(trailing)
        self.assertEqual(raised.exception.code, "ROW_TRAILING_CELL_NOT_ALLOWED")
        self.assertEqual(raised.exception.detail, "sheet=1. Area,row=2")

    def test_accepts_updated_valid_technician_codes_without_source_hash_lock(self) -> None:
        updated = _mutated_source(
            TECHNICIAN_SHEET_NAME,
            lambda sheet: setattr(sheet["A2"], "value", "AI999999"),
        )
        parsed = parse_atlanta_6area_workbook(updated)
        self.assertNotEqual(parsed.source_sha256, EXPECTED_SOURCE_SHA256)
        self.assertEqual(parsed.technicians[0].employee_code, "AI999999")
        bundle = build_atlanta_6area_bundle(
            updated, boundary_resolutions=_boundary_resolutions()
        )
        validated = validate_atlanta_6area_bundle(bundle.bundle_bytes)
        self.assertEqual(validated.manifest["source"]["sha256"], parsed.source_sha256)
        self.assertEqual(validated.plan_id, bundle.plan_id)

    def test_requires_exact_split_sheet_names_and_headers(self) -> None:
        with self.assertRaisesRegex(Atlanta6AreaPlanError, "HEADER_SCHEMA_INVALID"):
            parse_atlanta_6area_workbook(
                _mutated_source(
                    AREA_SHEET_NAME,
                    lambda sheet: setattr(sheet["A1"], "value", "PostalCode"),
                )
            )

        workbook = openpyxl.load_workbook(SOURCE)
        try:
            workbook[AREA_SHEET_NAME].title = "Sheet1"
            stream = io.BytesIO()
            workbook.save(stream)
        finally:
            workbook.close()
        with self.assertRaisesRegex(Atlanta6AreaPlanError, "WORKBOOK_SHEET_SCHEMA_INVALID"):
            parse_atlanta_6area_workbook(stream.getvalue())

    def test_required_boundary_decisions_fail_without_silent_defaults(self) -> None:
        cases = []
        missing_owner = _boundary_resolutions()
        missing_owner["30028"].pop("primary_region")
        cases.append(
            (
                "BOUNDARY_PRIMARY_REGION_INVALID",
                "postal=30028,field=primary_region",
                missing_owner,
            )
        )
        missing_overflow = _boundary_resolutions()
        missing_overflow["30028"].pop("allow_overflow")
        cases.append(
            (
                "BOUNDARY_ALLOW_OVERFLOW_INVALID",
                "postal=30028,field=allow_overflow",
                missing_overflow,
            )
        )
        for code, detail, decisions in cases:
            with self.subTest(code=code):
                with self.assertRaises(Atlanta6AreaPlanError) as raised:
                    build_atlanta_6area_bundle(
                        SOURCE, boundary_resolutions=decisions
                    )
                self.assertEqual(raised.exception.code, code)
                self.assertEqual(raised.exception.detail, detail)

    def test_rejects_non_zone_2_3_postal_conflict(self) -> None:
        def mutate(sheet) -> None:
            self.assertEqual(sheet["A121"].value, 30028)
            self.assertEqual(sheet["B121"].value, "Zone 3")
            sheet["B121"] = "Zone 4"

        with self.assertRaisesRegex(Atlanta6AreaPlanError, "POSTAL_MEMBERSHIP_CONFLICT"):
            parse_atlanta_6area_workbook(_mutated_source(AREA_SHEET_NAME, mutate))


if __name__ == "__main__":
    unittest.main()

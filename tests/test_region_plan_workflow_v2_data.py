import io
import unittest
from openpyxl import Workbook
from tools.data.region_plan_workflow_v2 import canonicalize_workbook, adopt_legacy_candidate, RegionPlanV2ValidationError

META={"subsidiary_id":"LGEAI","target_city_id":"los_angeles","source_city_id":"los_angeles","policy_version":"active_roster_area_type_fallback_region_soft/v1","technician_policy_mode":"active_roster_area_type_fallback_region_soft"}
def book(blank_assignment=False):
    wb=Workbook(); a=wb.active; a.title="Area"; a.append(["Territory","region_name","ZIPCode","Area Type","required_center_type"]); a.append(["R1","One","90001","DMS","DMS"])
    t=wb.create_sheet("Technician"); t.append(["Tech ID","Assignment","active"]); t.append(["AI1","R1","true"])
    if blank_assignment: t.append(["AI2","","true"])
    b=io.BytesIO(); wb.save(b); return b.getvalue()
class TestRegionPlanV2Data(unittest.TestCase):
 def test_aliases_hash_and_accounting(self):
  result=canonicalize_workbook(book(),META); m=result['manifest']; self.assertEqual(m['status'],'candidate'); self.assertTrue(m['plan_id'].startswith('rp2_los_angeles_')); self.assertEqual(m['row_accounting']['Area']['accepted_rows'],1)
 def test_area_plan_display_name_is_preserved_in_candidate_id(self):
  metadata={**META,"target_city_id":"Atlanta_GA","source_city_id":"Atlanta, GA","plan_display_name":"Atlanta_3area"}
  manifest=canonicalize_workbook(book(),metadata)['manifest']
  self.assertEqual(manifest['city_metadata']['target_city_id'],'Atlanta_GA')
  self.assertEqual(manifest['city_metadata']['plan_display_name'],'Atlanta_3area')
  self.assertTrue(manifest['plan_id'].startswith('rp2_atlanta_3area_'))
 def test_blank_assignment_is_rejected(self):
  result=canonicalize_workbook(book(True),META); self.assertEqual(result['manifest']['status'],'candidate'); self.assertEqual(result['rejects'][0]['error_code'],'TECHNICIAN_ASSIGNMENT_BLANK'); self.assertEqual(result['manifest']['excluded_rows']['TECHNICIAN_ASSIGNMENT_BLANK'],1)
 def test_actual_la_columns_are_accepted_without_pii(self):
  wb=Workbook(); a=wb.active; a.title='Area'; a.append(['ZIPCode','Territory','Area Type']); a.append(['90001','Region 1','DMS']); t=wb.create_sheet('Technician'); t.append(['Tech ID','Tech Name','Assignment']); t.append(['A1','Private Name','Region 1']); b=io.BytesIO(); wb.save(b)
  result=canonicalize_workbook(b.getvalue(),META); self.assertEqual(result['manifest']['areas'][0]['region_name'],'Region 1'); self.assertNotIn('Private Name',str(result['manifest']))
 def test_duplicate_and_overlap_preserve_row_accounting(self):
  wb=Workbook(); a=wb.active; a.title='Area'; a.append(['Territory','ZIPCode','Area Type']); a.append(['Region One','90001','DMS']); a.append(['Region Two','90001','DMS'])
  t=wb.create_sheet('Technician'); t.append(['Tech ID','Assignment']); t.append(['A','Region One']); t.append(['A','Region Two']); b=io.BytesIO(); wb.save(b)
  m=canonicalize_workbook(b.getvalue(),META)['manifest']; self.assertEqual(m['status'],'rejected'); self.assertEqual(m['row_accounting']['Technician']['accepted_rows']+m['row_accounting']['Technician']['rejected_rows'],2); self.assertIn('OVERLAP_POLICY_INVALID',m['plan_errors'])
 def test_generic_legacy_adoption_and_invalid_hash(self):
  files=tuple('abcdef'); x=adopt_legacy_candidate({'files':files,'area_count':413,'technician_count':54,'manifest_sha256':'a'*64,'source_sha256':'b'*64},expected_area_count=413,expected_technician_count=54,expected_filenames=files); self.assertEqual(x['technician_count'],54)
  with self.assertRaises(RegionPlanV2ValidationError): adopt_legacy_candidate({'files':files,'area_count':413,'technician_count':54,'manifest_sha256':'bad','source_sha256':'b'*64},expected_area_count=413,expected_technician_count=54,expected_filenames=files)
 def test_metadata_policy_mode_is_forced_even_when_workbook_has_older_mode(self):
  result=canonicalize_workbook(book(),META); self.assertEqual(result['manifest']['technicians'][0]['policy_mode'],META['technician_policy_mode'])
  wb=Workbook(); a=wb.active; a.title='Area'; a.append(['Territory','ZIPCode','Area Type']); a.append(['R1','90001','DMS']); t=wb.create_sheet('Technician'); t.append(['Tech ID','Assignment','policy_mode']); t.append(['A','R1','assigned_region_boundary_spillover']); b=io.BytesIO(); wb.save(b)
  result=canonicalize_workbook(b.getvalue(),META); self.assertEqual(result['rejects'],[]); self.assertEqual(result['manifest']['technicians'][0]['policy_mode'],META['technician_policy_mode'])
 def test_numbered_sheet_aliases_and_ambiguous_duplicate(self):
  wb=Workbook(); a=wb.active; a.title='1. Area'; a.append(['ZIPCode','Territory','Area Type']); a.append(['90001','Region 1','DMS']); t=wb.create_sheet('2. Technician'); t.append(['Tech ID','Tech Name','Assignment']); t.append(['A1','Private','Region 1']); b=io.BytesIO(); wb.save(b)
  self.assertEqual(canonicalize_workbook(b.getvalue(),META)['manifest']['status'],'candidate')
  wb.create_sheet('Area'); b=io.BytesIO(); wb.save(b)
  with self.assertRaises(RegionPlanV2ValidationError) as caught: canonicalize_workbook(b.getvalue(),META)
  self.assertEqual(caught.exception.code,'AMBIGUOUS_SHEET_ALIAS')
 def test_single_membership_overflow_and_partial_identity_are_fatal(self):
  wb=Workbook(); a=wb.active; a.title='Area'; a.append(['Territory','ZIPCode','Area Type','overflow_allowed','overflow_penalty_minutes']); a.append(['R1','90001','DMS','true','1'])
  t=wb.create_sheet('Technician'); t.append(['Tech ID','Assignment']); t.append(['','R1']); b=io.BytesIO(); wb.save(b)
  result=canonicalize_workbook(b.getvalue(),META)
  self.assertEqual('rejected',result['manifest']['status']); self.assertIn('OVERLAP_POLICY_INVALID',result['manifest']['plan_errors']); self.assertEqual('TECHNICIAN_ROW_INVALID',result['rejects'][0]['error_code'])

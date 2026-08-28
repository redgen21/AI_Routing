"""Pure preprocessing contract for the reviewed Atlanta six-area workbook.

The source workbook has two independent, explicitly named sheets: postal/
territory memberships on ``1. Area`` and the technician roster on
``2. Technician``.  Rows in the two sheets are not related by position.  This
module validates that immutable source contract and returns deterministic
artifacts in memory; it never writes files or connects to a database.
"""
from __future__ import annotations

import csv
import hashlib
import io
import json
import re
import zipfile
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from types import MappingProxyType
from typing import Any, Mapping

import openpyxl
from openpyxl.comments import Comment
from openpyxl.worksheet.datavalidation import DataValidation


PLAN_ID = "atlanta_6area_new_atl_buckets_20260721_v2"
STRATEGIC_CITY_NAME = "Atlanta_6area"
SOURCE_FILE_NAME = "New ATL Buckets.xlsx"
AREA_SHEET_NAME = "1. Area"
TECHNICIAN_SHEET_NAME = "2. Technician"
MANIFEST_SCHEMA = "atlanta-6area-plan-bundle/v2"
POLICY_VERSION = "own_region_with_approved_boundary_overflow/v2"
FIXED_REGION_SCHEMA = "fixed-region-postal/v1"
BOUNDARY_POLICY_SCHEMA = "fixed-region-boundary-policy/v1"
TECHNICIAN_POLICY_SCHEMA = "fixed-region-technician-policy/v2"
BOUNDARY_PENALTY_COST = 4500
CANONICAL_AREA_TYPE = "DMS"
_VERSIONED_PLAN_ID_RE = re.compile(r"^atlanta_6area_v2_([0-9a-f]{64})$")
_SOURCE_TECHNICIAN_MASTER_CONTEXT = MappingProxyType(
    {"subsidiary_name": "LGEAI", "strategic_city_name": "Atlanta, GA"}
)
_TARGET_TECHNICIAN_MASTER_CONTEXT = MappingProxyType(
    {"subsidiary_name": "LGEAI", "strategic_city_name": STRATEGIC_CITY_NAME}
)
_PRIVACY_CLASSIFICATION = "internal_pii_redacted"

ZONE_TO_SEQ: Mapping[str, int] = MappingProxyType(
    {
        "Zone 1": 1,
        "Zone 2": 2,
        "Zone 3": 3,
        "Zone 4": 4,
        "Zone 5": 5,
        "ATL Outer Area": 6,
    }
)
EXPECTED_TECH_IDS = frozenset(
    {
        "AI005576",
        "AI102087",
        "AI102315",
        "AI102448",
        "AI102608",
        "AI102961",
        "AI102977",
        "AI103128",
        "AI103146",
        "AI103261",
        "AI103264",
        "AI103317",
        "AI105115",
        "AI105116",
    }
)
# The historic workbook IDs remain available as a regression fixture only.
# Current canonical uploads intentionally accept any 14 distinct, syntactically
# valid AI codes; the transactionally-read technician master is the authority
# for active/source-master eligibility.
EXPECTED_TECHNICIAN_ROWS = 14
EXPECTED_AMBIGUOUS_POSTALS = ("30028", "30040", "30041", "30107")
EXPECTED_MEMBERSHIP_ROWS = 301
EXPECTED_UNIQUE_POSTALS = 297
# This is the immutable source workbook approved for the six-area candidate.
# The bundle deliberately carries the source checksum rather than the workbook
# itself, so consumers can validate provenance without exposing its roster.
EXPECTED_SOURCE_SHA256 = "19cd5a42ef3f09e120dd84b26cc202deddc77aedcd08dbfc015d1a4144aeaedb"

FIXED_REGION_FILENAME = f"fixed_region_postal_atlanta_6area_{PLAN_ID}.csv"
BOUNDARY_POLICY_FILENAME = f"atlanta_6area_boundary_policy_{PLAN_ID}.csv"
TECHNICIAN_POLICY_FILENAME = f"atlanta_6area_technician_policy_{PLAN_ID}.csv"
MANIFEST_FILENAME = f"atlanta_6area_plan_{PLAN_ID}.json"

_AREA_HEADERS = ("ZIPCode", "Territory")
_TECHNICIAN_HEADERS = ("Tech ID", "Tech Name", "Assignment")
_TECH_ID_RE = re.compile(r"^AI\d{6}$")
_POSTAL_RE = re.compile(r"^\d{5}$")
_TEMPLATE_TIMESTAMP = datetime(1980, 1, 1, 0, 0, 0)
_FIXED_REGION_COLUMNS = (
    "POSTAL_CODE",
    "STRATEGIC_CITY_NAME",
    "region_id",
    "region_seq",
    "AREA_NAME",
    "new_region_name",
    "area_type",
)
_BOUNDARY_POLICY_COLUMNS = (
    "plan_id",
    "STRATEGIC_CITY_NAME",
    "POSTAL_CODE",
    "primary_region_id",
    "primary_region_seq",
    "alternate_region_id",
    "alternate_region_seq",
    "policy_mode",
    "penalty_cost",
)
_TECHNICIAN_POLICY_COLUMNS = (
    "plan_id",
    "STRATEGIC_CITY_NAME",
    "SVC_ENGINEER_CODE",
    "assigned_region_id",
    "assigned_region_seq",
    "assigned_region_name",
    "policy_mode",
)


class Atlanta6AreaPlanError(ValueError):
    """Stable, non-sensitive validation failure."""

    def __init__(self, code: str, detail: str = ""):
        self.code = code
        self.detail = detail
        message = code if not detail else f"{code}: {detail}"
        super().__init__(message)


@dataclass(frozen=True)
class PostalMembership:
    source_row: int
    postal_code: str
    territory: str


@dataclass(frozen=True)
class TechnicianAssignment:
    source_row: int
    employee_code: str
    territory: str


@dataclass(frozen=True)
class ParsedAtlanta6AreaWorkbook:
    source_sha256: str
    memberships: tuple[PostalMembership, ...]
    technicians: tuple[TechnicianAssignment, ...]
    postal_territories: Mapping[str, tuple[str, ...]]
    ambiguous_postals: tuple[str, ...]
    territory_membership_counts: Mapping[str, int]


@dataclass(frozen=True)
class Atlanta6AreaPreview:
    promotable: bool
    approval_status: str
    summary: Mapping[str, Any]

    def as_dict(self) -> dict[str, Any]:
        return {
            "promotable": self.promotable,
            "approval_status": self.approval_status,
            "summary": dict(self.summary),
        }


@dataclass(frozen=True)
class Atlanta6AreaBundle:
    plan_id: str
    promotable: bool
    approval_status: str
    summary: Mapping[str, Any]
    manifest: Mapping[str, Any]
    artifacts: Mapping[str, bytes]
    bundle_bytes: bytes

    def as_dict(self) -> dict[str, Any]:
        return {
            "plan_id": self.plan_id,
            "promotable": self.promotable,
            "approval_status": self.approval_status,
            "summary": dict(self.summary),
            "manifest": dict(self.manifest),
            "artifact_names": sorted(self.artifacts),
            "bundle_size_bytes": len(self.bundle_bytes),
        }


def _source_bytes(source: bytes | bytearray | memoryview | Path | str) -> bytes:
    if isinstance(source, bytes):
        payload = source
    elif isinstance(source, (bytearray, memoryview)):
        payload = bytes(source)
    elif isinstance(source, (Path, str)):
        path = Path(source)
        if not path.is_file():
            raise Atlanta6AreaPlanError("SOURCE_NOT_FOUND")
        payload = path.read_bytes()
    else:
        raise Atlanta6AreaPlanError("SOURCE_TYPE_NOT_ALLOWED")
    if not payload:
        raise Atlanta6AreaPlanError("SOURCE_EMPTY")
    return payload


def _source_metadata(source_sha256: str) -> dict[str, Any]:
    """Return non-PII lineage for the two independently accounted sheets."""

    return {
        "file_name": SOURCE_FILE_NAME,
        "sha256": source_sha256,
        "sheets": [
            {
                "role": "membership",
                "sheet_name": AREA_SHEET_NAME,
                "headers": list(_AREA_HEADERS),
                "input_rows": EXPECTED_MEMBERSHIP_ROWS,
            },
            {
                "role": "technician",
                "sheet_name": TECHNICIAN_SHEET_NAME,
                "headers": list(_TECHNICIAN_HEADERS),
                "input_rows": EXPECTED_TECHNICIAN_ROWS,
            },
        ],
    }


def _clean(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize_postal(value: object, *, source_row: int) -> str:
    context = f"row={source_row},field=ZIPCode"
    if value is None or isinstance(value, bool):
        raise Atlanta6AreaPlanError("ZIP_INVALID", context)
    if isinstance(value, int):
        text = str(value)
    elif isinstance(value, float):
        if not value.is_integer():
            raise Atlanta6AreaPlanError("ZIP_INVALID", context)
        text = str(int(value))
    else:
        text = str(value).strip()
        if re.fullmatch(r"\d{1,5}\.0", text):
            text = text[:-2]
    if not re.fullmatch(r"\d{1,5}", text):
        raise Atlanta6AreaPlanError("ZIP_INVALID", context)
    normalized = text.zfill(5)
    if not re.fullmatch(r"\d{5}", normalized):
        raise Atlanta6AreaPlanError("ZIP_INVALID", context)
    return normalized


def _load_sheets(payload: bytes):
    try:
        workbook = openpyxl.load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
    except Exception as exc:
        raise Atlanta6AreaPlanError("WORKBOOK_INVALID") from exc
    expected_sheet_names = (AREA_SHEET_NAME, TECHNICIAN_SHEET_NAME)
    if tuple(workbook.sheetnames) != expected_sheet_names:
        workbook.close()
        raise Atlanta6AreaPlanError(
            "WORKBOOK_SHEET_SCHEMA_INVALID",
            f"expected={','.join(expected_sheet_names)}",
        )
    return workbook, workbook[AREA_SHEET_NAME], workbook[TECHNICIAN_SHEET_NAME]


def build_atlanta_6area_template_bytes() -> bytes:
    """Return a deterministic blank workbook for the canonical source contract.

    The template deliberately contains only the two source sheets and their
    exact headers.  It is a download payload, not a candidate: blank data rows
    cannot satisfy the fixed 301/297/14/4 source acceptance checks.
    """

    workbook = openpyxl.Workbook()
    try:
        area_sheet = workbook.active
        area_sheet.title = AREA_SHEET_NAME
        technician_sheet = workbook.create_sheet(TECHNICIAN_SHEET_NAME)
        workbook.properties.creator = "VRP Routing"
        workbook.properties.lastModifiedBy = "VRP Routing"
        workbook.properties.created = _TEMPLATE_TIMESTAMP
        workbook.properties.modified = _TEMPLATE_TIMESTAMP

        for sheet, headers in (
            (area_sheet, _AREA_HEADERS),
            (technician_sheet, _TECHNICIAN_HEADERS),
        ):
            sheet.append(headers)
            sheet.freeze_panes = "A2"
            for cell in sheet[1]:
                cell.comment = Comment(
                    "Exact canonical source header; do not rename or add columns.",
                    "VRP Routing",
                )

        area_sheet["A1"].comment = Comment(
            "USPS ZIP5. Numeric values are normalized to five digits during validation.",
            "VRP Routing",
        )
        area_sheet["B1"].comment = Comment(
            "Allowed values: Zone 1 through Zone 5, or ATL Outer Area.",
            "VRP Routing",
        )
        technician_sheet["A1"].comment = Comment(
            "Approved technician identifier, e.g. AI005576.", "VRP Routing"
        )
        technician_sheet["B1"].comment = Comment(
            "Source name is required for validation and is never included in bundle outputs.",
            "VRP Routing",
        )
        technician_sheet["C1"].comment = Comment(
            "Allowed values: Zone 1 through Zone 5, or ATL Outer Area.",
            "VRP Routing",
        )
        allowed_territories = '"Zone 1,Zone 2,Zone 3,Zone 4,Zone 5,ATL Outer Area"'
        for sheet, range_ref in ((area_sheet, "B2:B10000"), (technician_sheet, "C2:C10000")):
            validation = DataValidation(type="list", formula1=allowed_territories)
            validation.error = "Select one of the canonical Atlanta territories."
            validation.errorTitle = "Territory not allowed"
            validation.prompt = "Choose a canonical territory."
            validation.promptTitle = "Territory"
            validation.showErrorMessage = True
            validation.showInputMessage = True
            sheet.add_data_validation(validation)
            validation.add(range_ref)

        stream = io.BytesIO()
        workbook.save(stream)
    finally:
        workbook.close()

    # ``openpyxl`` timestamps ZIP entries at write time.  Canonicalizing that
    # envelope makes the download bytes reproducible without altering workbook
    # semantics or adding a hidden helper sheet.
    with zipfile.ZipFile(io.BytesIO(stream.getvalue()), "r") as archive:
        entries = {name: archive.read(name) for name in archive.namelist()}
    return _deterministic_zip(entries)


def get_atlanta_6area_template_bytes() -> bytes:
    """Download-bytes API for callers that do not need a filesystem artifact."""

    return build_atlanta_6area_template_bytes()


def get_atlanta_6area_workbook_template() -> bytes:
    """Compatibility download API for the canonical blank workbook."""

    return build_atlanta_6area_template_bytes()


def _sheet_rows(sheet: Any, headers: tuple[str, ...], *, sheet_name: str):
    rows = sheet.iter_rows(values_only=True)
    header_row = next(rows, None)
    if header_row is None:
        raise Atlanta6AreaPlanError("HEADER_MISSING", f"sheet={sheet_name},row=1")
    actual_headers = tuple(_clean(value) for value in header_row[: len(headers)])
    if actual_headers != headers or any(_clean(value) for value in header_row[len(headers) :]):
        raise Atlanta6AreaPlanError(
            "HEADER_SCHEMA_INVALID", f"sheet={sheet_name},row=1"
        )
    return rows


def _canonical_row_values(
    row: tuple[object, ...],
    headers: tuple[str, ...],
    *,
    sheet_name: str,
    source_row: int,
) -> tuple[object, ...]:
    """Reject values outside the exact source columns; blank rows stay valid."""

    trailing = row[len(headers):]
    if any(_clean(value) for value in trailing):
        raise Atlanta6AreaPlanError(
            "ROW_TRAILING_CELL_NOT_ALLOWED",
            f"sheet={sheet_name},row={source_row}",
        )
    return tuple(row[: len(headers)]) + (None,) * max(0, len(headers) - len(row))


def parse_atlanta_6area_workbook(
    source: bytes | bytearray | memoryview | Path | str,
) -> ParsedAtlanta6AreaWorkbook:
    """Parse and validate the fixed New ATL Buckets source contract."""

    payload = _source_bytes(source)
    workbook, area_sheet, technician_sheet = _load_sheets(payload)
    try:
        area_rows = _sheet_rows(area_sheet, _AREA_HEADERS, sheet_name=AREA_SHEET_NAME)
        technician_rows = _sheet_rows(
            technician_sheet, _TECHNICIAN_HEADERS, sheet_name=TECHNICIAN_SHEET_NAME
        )

        memberships: list[PostalMembership] = []
        technicians: list[TechnicianAssignment] = []
        seen_memberships: set[tuple[str, str]] = set()
        seen_tech_ids: set[str] = set()

        for source_row, row in enumerate(area_rows, start=2):
            raw_postal, raw_territory_value = _canonical_row_values(
                row, _AREA_HEADERS, sheet_name=AREA_SHEET_NAME, source_row=source_row
            )
            raw_territory = _clean(raw_territory_value)
            has_postal = raw_postal is not None and _clean(raw_postal) != ""
            if has_postal or raw_territory:
                if not has_postal:
                    raise Atlanta6AreaPlanError(
                        "POSTAL_MEMBERSHIP_INCOMPLETE",
                        f"row={source_row},field=ZIPCode",
                    )
                if not raw_territory:
                    raise Atlanta6AreaPlanError(
                        "POSTAL_MEMBERSHIP_INCOMPLETE",
                        f"row={source_row},field=Territory",
                    )
                if raw_territory not in ZONE_TO_SEQ:
                    raise Atlanta6AreaPlanError(
                        "TERRITORY_INVALID", f"row={source_row},field=Territory"
                    )
                postal_code = _normalize_postal(raw_postal, source_row=source_row)
                membership_key = (postal_code, raw_territory)
                if membership_key in seen_memberships:
                    raise Atlanta6AreaPlanError(
                        "POSTAL_MEMBERSHIP_DUPLICATE",
                        f"row={source_row},fields=ZIPCode,Territory",
                    )
                seen_memberships.add(membership_key)
                memberships.append(PostalMembership(source_row, postal_code, raw_territory))

        for source_row, row in enumerate(technician_rows, start=2):
            employee_code_value, employee_name_value, assignment_value = _canonical_row_values(
                row,
                _TECHNICIAN_HEADERS,
                sheet_name=TECHNICIAN_SHEET_NAME,
                source_row=source_row,
            )
            employee_code = _clean(employee_code_value).upper()
            employee_name = _clean(employee_name_value)
            assignment = _clean(assignment_value)
            has_any_tech = bool(employee_code or employee_name or assignment)
            if has_any_tech:
                if not employee_code:
                    raise Atlanta6AreaPlanError(
                        "TECH_ID_MISSING", f"row={source_row},field=Tech ID"
                    )
                if not _TECH_ID_RE.fullmatch(employee_code):
                    raise Atlanta6AreaPlanError(
                        "TECH_ID_UNKNOWN", f"row={source_row},field=Tech ID"
                    )
                if employee_code in seen_tech_ids:
                    raise Atlanta6AreaPlanError(
                        "TECH_ID_DUPLICATE", f"row={source_row},field=Tech ID"
                    )
                if not employee_name:
                    raise Atlanta6AreaPlanError(
                        "TECH_NAME_MISSING", f"row={source_row},field=Tech Name"
                    )
                if assignment not in ZONE_TO_SEQ:
                    raise Atlanta6AreaPlanError(
                        "TECH_ASSIGNMENT_INVALID",
                        f"row={source_row},field=Assignment",
                    )
                seen_tech_ids.add(employee_code)
                technicians.append(
                    TechnicianAssignment(source_row, employee_code, assignment)
                )
    finally:
        workbook.close()

    if len(memberships) != EXPECTED_MEMBERSHIP_ROWS:
        raise Atlanta6AreaPlanError(
            "MEMBERSHIP_ROW_COUNT_INVALID",
            f"expected={EXPECTED_MEMBERSHIP_ROWS},actual={len(memberships)}",
        )

    postal_groups: dict[str, set[str]] = defaultdict(set)
    for membership in memberships:
        postal_groups[membership.postal_code].add(membership.territory)
    if len(postal_groups) != EXPECTED_UNIQUE_POSTALS:
        raise Atlanta6AreaPlanError(
            "UNIQUE_POSTAL_COUNT_INVALID",
            f"expected={EXPECTED_UNIQUE_POSTALS},actual={len(postal_groups)}",
        )

    ambiguous_postals: list[str] = []
    for postal_code, territories in sorted(postal_groups.items()):
        if len(territories) <= 1:
            continue
        if territories != {"Zone 2", "Zone 3"}:
            raise Atlanta6AreaPlanError("POSTAL_MEMBERSHIP_CONFLICT", postal_code)
        ambiguous_postals.append(postal_code)
    if tuple(ambiguous_postals) != EXPECTED_AMBIGUOUS_POSTALS:
        raise Atlanta6AreaPlanError(
            "AMBIGUOUS_POSTALS_INVALID", ",".join(ambiguous_postals)
        )

    if len(technicians) != EXPECTED_TECHNICIAN_ROWS:
        raise Atlanta6AreaPlanError(
            "TECHNICIAN_ROW_COUNT_INVALID",
            f"expected={EXPECTED_TECHNICIAN_ROWS},actual={len(technicians)}",
        )

    postal_territories = MappingProxyType(
        {
            postal_code: tuple(sorted(territories, key=lambda name: ZONE_TO_SEQ[name]))
            for postal_code, territories in sorted(postal_groups.items())
        }
    )
    territory_counts = Counter(membership.territory for membership in memberships)
    return ParsedAtlanta6AreaWorkbook(
        source_sha256=hashlib.sha256(payload).hexdigest(),
        memberships=tuple(memberships),
        technicians=tuple(technicians),
        postal_territories=postal_territories,
        ambiguous_postals=tuple(ambiguous_postals),
        territory_membership_counts=MappingProxyType(
            {territory: int(territory_counts[territory]) for territory in ZONE_TO_SEQ}
        ),
    )


def _safe_summary(
    parsed: ParsedAtlanta6AreaWorkbook, *, plan_id: str = PLAN_ID
) -> dict[str, Any]:
    return {
        "plan_id": plan_id,
        "strategic_city_name": STRATEGIC_CITY_NAME,
        "source_sha256": parsed.source_sha256,
        "membership_input_rows": len(parsed.memberships),
        "membership_accepted_rows": len(parsed.memberships),
        "membership_rejected_rows": 0,
        "unique_postal_count": len(parsed.postal_territories),
        "ambiguous_postal_count": len(parsed.ambiguous_postals),
        "ambiguous_postals": list(parsed.ambiguous_postals),
        "technician_input_rows": len(parsed.technicians),
        "technician_accepted_rows": len(parsed.technicians),
        "technician_rejected_rows": 0,
        "territory_membership_counts": dict(parsed.territory_membership_counts),
    }


def preview_atlanta_6area_plan(
    source: bytes | bytearray | memoryview | Path | str,
) -> Atlanta6AreaPreview:
    """Return a PII-safe validation preview without resolving the boundary."""

    parsed = parse_atlanta_6area_workbook(source)
    return Atlanta6AreaPreview(
        promotable=False,
        approval_status="pending_boundary_resolutions",
        summary=MappingProxyType(_safe_summary(parsed)),
    )


def _region_id(territory: str) -> str:
    return f"atlanta_6area_r{ZONE_TO_SEQ[territory]:02d}"


def _region_name(territory: str) -> str:
    return f"Atlanta_6area {territory}"


def _csv_bytes(fieldnames: tuple[str, ...], rows: list[dict[str, object]]) -> bytes:
    stream = io.StringIO(newline="")
    writer = csv.DictWriter(stream, fieldnames=fieldnames, lineterminator="\n")
    writer.writeheader()
    writer.writerows(rows)
    return stream.getvalue().encode("utf-8-sig")


def _canonical_artifact_contract() -> dict[str, dict[str, Any]]:
    """Describe stable keys, units, and null policy for DB-input projections."""

    return {
        FIXED_REGION_FILENAME: {
            "schema": FIXED_REGION_SCHEMA,
            "logical_record": "region_postal",
            "columns": list(_FIXED_REGION_COLUMNS),
            "key_columns": ["POSTAL_CODE"],
            "units": {"POSTAL_CODE": "USPS_ZIP5", "region_seq": "ordinal"},
            "null_policy": {"area_type": "non_null_constant_DMS"},
        },
        BOUNDARY_POLICY_FILENAME: {
            "schema": BOUNDARY_POLICY_SCHEMA,
            "logical_record": "boundary_policy",
            "columns": list(_BOUNDARY_POLICY_COLUMNS),
            "key_columns": ["plan_id", "POSTAL_CODE"],
            "units": {
                "primary_region_seq": "ordinal",
                "alternate_region_seq": "ordinal",
                "penalty_cost": "solver_objective_cost",
            },
            "null_policy": {},
        },
        TECHNICIAN_POLICY_FILENAME: {
            "schema": TECHNICIAN_POLICY_SCHEMA,
            "logical_record": "technician_region_policy",
            "columns": list(_TECHNICIAN_POLICY_COLUMNS),
            "key_columns": ["plan_id", "SVC_ENGINEER_CODE"],
            "units": {"assigned_region_seq": "ordinal"},
            "null_policy": {},
            "privacy": {
                "technician_names_redacted": True,
                "prohibited_columns": ["SVC_ENGINEER_NAME"],
            },
        },
    }


def _fixed_region_bytes(
    parsed: ParsedAtlanta6AreaWorkbook,
    boundary_resolutions: Mapping[str, Mapping[str, Any]],
) -> bytes:
    rows: list[dict[str, object]] = []
    for postal_code, territories in parsed.postal_territories.items():
        territory = (
            str(boundary_resolutions[postal_code]["primary_region"])
            if len(territories) > 1
            else territories[0]
        )
        rows.append(
            {
                "POSTAL_CODE": postal_code,
                "STRATEGIC_CITY_NAME": STRATEGIC_CITY_NAME,
                "region_id": _region_id(territory),
                "region_seq": ZONE_TO_SEQ[territory],
                "AREA_NAME": territory,
                "new_region_name": _region_name(territory),
                "area_type": CANONICAL_AREA_TYPE,
            }
        )
    return _csv_bytes(
        (
            "POSTAL_CODE",
            "STRATEGIC_CITY_NAME",
            "region_id",
            "region_seq",
            "AREA_NAME",
            "new_region_name",
            "area_type",
        ),
        rows,
    )


def _boundary_policy_bytes(
    parsed: ParsedAtlanta6AreaWorkbook,
    boundary_resolutions: Mapping[str, Mapping[str, Any]],
    *,
    plan_id: str,
) -> bytes:
    rows: list[dict[str, object]] = []
    for postal_code in parsed.ambiguous_postals:
        decision = boundary_resolutions[postal_code]
        if decision["allow_overflow"] is not True:
            continue
        primary = str(decision["primary_region"])
        alternate = "Zone 3" if primary == "Zone 2" else "Zone 2"
        rows.append(
            {
                "plan_id": plan_id,
                "STRATEGIC_CITY_NAME": STRATEGIC_CITY_NAME,
                "POSTAL_CODE": postal_code,
                "primary_region_id": _region_id(primary),
                "primary_region_seq": ZONE_TO_SEQ[primary],
                "alternate_region_id": _region_id(alternate),
                "alternate_region_seq": ZONE_TO_SEQ[alternate],
                "policy_mode": "boundary_primary_alternate_soft",
                "penalty_cost": BOUNDARY_PENALTY_COST,
            }
        )
    return _csv_bytes(
        (
            "plan_id",
            "STRATEGIC_CITY_NAME",
            "POSTAL_CODE",
            "primary_region_id",
            "primary_region_seq",
            "alternate_region_id",
            "alternate_region_seq",
            "policy_mode",
            "penalty_cost",
        ),
        rows,
    )


def _technician_policy_bytes(
    parsed: ParsedAtlanta6AreaWorkbook, *, plan_id: str
) -> bytes:
    rows = []
    for technician in sorted(parsed.technicians, key=lambda item: item.employee_code):
        rows.append(
            {
                "plan_id": plan_id,
                "STRATEGIC_CITY_NAME": STRATEGIC_CITY_NAME,
                "SVC_ENGINEER_CODE": technician.employee_code,
                "assigned_region_id": _region_id(technician.territory),
                "assigned_region_seq": ZONE_TO_SEQ[technician.territory],
                "assigned_region_name": _region_name(technician.territory),
                "policy_mode": "assigned_region_boundary_spillover",
            }
        )
    return _csv_bytes(
        (
            "plan_id",
            "STRATEGIC_CITY_NAME",
            "SVC_ENGINEER_CODE",
            "assigned_region_id",
            "assigned_region_seq",
            "assigned_region_name",
            "policy_mode",
        ),
        rows,
    )


def _artifact_metadata(payload: bytes, row_count: int) -> dict[str, Any]:
    return {
        "sha256": hashlib.sha256(payload).hexdigest(),
        "size_bytes": len(payload),
        "row_count": row_count,
    }


def _deterministic_zip(artifacts: Mapping[str, bytes]) -> bytes:
    output = io.BytesIO()
    with zipfile.ZipFile(output, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=9) as archive:
        for name in sorted(artifacts):
            info = zipfile.ZipInfo(name, date_time=(1980, 1, 1, 0, 0, 0))
            info.compress_type = zipfile.ZIP_DEFLATED
            info.external_attr = 0o600 << 16
            archive.writestr(info, artifacts[name], compress_type=zipfile.ZIP_DEFLATED, compresslevel=9)
    return output.getvalue()


def _normalize_boundary_resolutions(
    value: Mapping[str, Mapping[str, Any]],
) -> Mapping[str, Mapping[str, Any]]:
    if not isinstance(value, Mapping):
        raise Atlanta6AreaPlanError(
            "BOUNDARY_RESOLUTIONS_REQUIRED", "field=boundary_resolutions"
        )
    submitted = {str(postal).strip(): decision for postal, decision in value.items()}
    expected = set(EXPECTED_AMBIGUOUS_POSTALS)
    missing = sorted(expected - set(submitted))
    unknown = sorted(set(submitted) - expected)
    if missing:
        raise Atlanta6AreaPlanError(
            "BOUNDARY_RESOLUTION_MISSING",
            f"field=boundary_resolutions,postal_codes={','.join(missing)}",
        )
    if unknown:
        raise Atlanta6AreaPlanError("BOUNDARY_RESOLUTION_UNKNOWN", ",".join(unknown))

    normalized: dict[str, Mapping[str, Any]] = {}
    for postal_code in EXPECTED_AMBIGUOUS_POSTALS:
        raw = submitted[postal_code]
        if not isinstance(raw, Mapping):
            raise Atlanta6AreaPlanError(
                "BOUNDARY_RESOLUTION_INVALID",
                f"postal={postal_code},field=boundary_resolution",
            )
        primary = _clean(raw.get("primary_region"))
        if primary not in {"Zone 2", "Zone 3"}:
            raise Atlanta6AreaPlanError(
                "BOUNDARY_PRIMARY_REGION_INVALID",
                f"postal={postal_code},field=primary_region",
            )
        allow_overflow = raw.get("allow_overflow")
        if not isinstance(allow_overflow, bool):
            raise Atlanta6AreaPlanError(
                "BOUNDARY_ALLOW_OVERFLOW_INVALID",
                f"postal={postal_code},field=allow_overflow",
            )
        rationale_value = raw.get("rationale", "")
        if rationale_value is None:
            rationale = ""
        elif not isinstance(rationale_value, str):
            raise Atlanta6AreaPlanError(
                "BOUNDARY_RATIONALE_INVALID",
                f"postal={postal_code},field=rationale",
            )
        else:
            rationale = " ".join(rationale_value.split())
        if len(rationale) > 500:
            raise Atlanta6AreaPlanError(
                "BOUNDARY_RATIONALE_INVALID",
                f"postal={postal_code},field=rationale",
            )
        alternate = "Zone 3" if primary == "Zone 2" else "Zone 2"
        normalized[postal_code] = MappingProxyType(
            {
                "primary_region": primary,
                "primary_region_seq": ZONE_TO_SEQ[primary],
                "alternate_region": alternate,
                "alternate_region_seq": ZONE_TO_SEQ[alternate],
                "allow_overflow": allow_overflow,
                "penalty_cost": BOUNDARY_PENALTY_COST if allow_overflow else None,
                "rationale": rationale,
            }
        )
    return MappingProxyType(normalized)


def derive_atlanta_6area_plan_identity(
    source_sha256: str,
    boundary_resolutions: Mapping[str, Mapping[str, Any]],
) -> tuple[str, str]:
    """Return the immutable digest and versioned ID for one resolved candidate."""

    source_digest = str(source_sha256).strip().lower()
    if not re.fullmatch(r"[0-9a-f]{64}", source_digest):
        raise Atlanta6AreaPlanError("SOURCE_CHECKSUM_INVALID")
    resolutions = _normalize_boundary_resolutions(boundary_resolutions)
    identity = {
        "source_sha256": source_digest,
        "boundary_resolutions": [
            {"postal_code": postal_code, **dict(resolutions[postal_code])}
            for postal_code in sorted(resolutions)
        ],
        "policy_version": POLICY_VERSION,
        "schema_version": MANIFEST_SCHEMA,
        "source_technician_master_context": dict(_SOURCE_TECHNICIAN_MASTER_CONTEXT),
        "target_technician_master_context": dict(_TARGET_TECHNICIAN_MASTER_CONTEXT),
        "technician_names_redacted": True,
    }
    digest = hashlib.sha256(
        json.dumps(identity, ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode(
            "utf-8"
        )
    ).hexdigest()
    return digest, f"atlanta_6area_v2_{digest}"


def is_allowed_atlanta_6area_plan_id(value: object) -> bool:
    plan_id = str(value).strip()
    return plan_id == PLAN_ID or _VERSIONED_PLAN_ID_RE.fullmatch(plan_id) is not None


def build_atlanta_6area_bundle(
    source: bytes | bytearray | memoryview | Path | str,
    *,
    boundary_resolutions: Mapping[str, Mapping[str, Any]],
) -> Atlanta6AreaBundle:
    """Build a deterministic bundle after four explicit ZIP-level decisions."""

    resolutions = _normalize_boundary_resolutions(boundary_resolutions)
    parsed = parse_atlanta_6area_workbook(source)
    resolution_digest, plan_id = derive_atlanta_6area_plan_identity(
        parsed.source_sha256, resolutions
    )
    fixed_region = _fixed_region_bytes(parsed, resolutions)
    boundary_policy = _boundary_policy_bytes(parsed, resolutions, plan_id=plan_id)
    technician_policy = _technician_policy_bytes(parsed, plan_id=plan_id)
    overflow_count = sum(
        1 for decision in resolutions.values() if decision["allow_overflow"] is True
    )
    artifact_payloads: dict[str, bytes] = {
        FIXED_REGION_FILENAME: fixed_region,
        BOUNDARY_POLICY_FILENAME: boundary_policy,
        TECHNICIAN_POLICY_FILENAME: technician_policy,
    }
    artifact_manifest = {
        FIXED_REGION_FILENAME: _artifact_metadata(fixed_region, EXPECTED_UNIQUE_POSTALS),
        BOUNDARY_POLICY_FILENAME: _artifact_metadata(
            boundary_policy, overflow_count
        ),
        TECHNICIAN_POLICY_FILENAME: _artifact_metadata(
            technician_policy, len(parsed.technicians)
        ),
    }
    summary = _safe_summary(parsed, plan_id=plan_id)
    summary.update(
        {
            "fixed_region_rows": EXPECTED_UNIQUE_POSTALS,
            "canonical_region_count": len(ZONE_TO_SEQ),
            "canonical_postal_count": EXPECTED_UNIQUE_POSTALS,
            "canonical_technician_count": len(parsed.technicians),
            "canonical_boundary_resolution_count": len(resolutions),
            "canonical_boundary_policy_count": overflow_count,
            "boundary_resolution_rows": len(resolutions),
            "boundary_policy_rows": overflow_count,
            "technician_policy_rows": len(parsed.technicians),
            "boundary_overflow_enabled_count": overflow_count,
            "boundary_penalty_cost": BOUNDARY_PENALTY_COST,
        }
    )
    manifest: dict[str, Any] = {
        "schema": MANIFEST_SCHEMA,
        "schema_version": MANIFEST_SCHEMA,
        "policy_version": POLICY_VERSION,
        "plan_id": plan_id,
        "resolution_digest": resolution_digest,
        "strategic_city_name": STRATEGIC_CITY_NAME,
        "approval_status": "resolved_for_development_verification",
        "promotable": False,
        "verification_only": True,
        "lifecycle_stage": "resolved_candidate",
        "privacy_classification": _PRIVACY_CLASSIFICATION,
        "technician_names_redacted": True,
        "source_technician_master_context": dict(_SOURCE_TECHNICIAN_MASTER_CONTEXT),
        "target_technician_master_context": dict(_TARGET_TECHNICIAN_MASTER_CONTEXT),
        "source": _source_metadata(parsed.source_sha256),
        "parents": [
            {
                "kind": "source_workbook",
                "file_name": SOURCE_FILE_NAME,
                "sha256": parsed.source_sha256,
            }
        ],
        "canonical_artifacts": _canonical_artifact_contract(),
        "boundary_resolutions": {
            postal_code: dict(decision)
            for postal_code, decision in resolutions.items()
        },
        "row_accounting": {
            "membership": {
                "input": len(parsed.memberships),
                "accepted": len(parsed.memberships),
                "rejected": 0,
                "unique_postals": len(parsed.postal_territories),
                "ambiguous_postals": len(parsed.ambiguous_postals),
            },
            "technician": {
                "input": len(parsed.technicians),
                "accepted": len(parsed.technicians),
                "rejected": 0,
            },
        },
        "zone_to_region_seq": dict(ZONE_TO_SEQ),
        "artifacts": artifact_manifest,
    }
    manifest_bytes = (
        json.dumps(manifest, ensure_ascii=False, indent=2, sort_keys=True) + "\n"
    ).encode("utf-8")
    artifact_payloads[MANIFEST_FILENAME] = manifest_bytes
    frozen_artifacts = MappingProxyType(dict(artifact_payloads))
    return Atlanta6AreaBundle(
        plan_id=plan_id,
        promotable=False,
        approval_status="resolved_for_development_verification",
        summary=MappingProxyType(summary),
        manifest=MappingProxyType(manifest),
        artifacts=frozen_artifacts,
        bundle_bytes=_deterministic_zip(frozen_artifacts),
    )


def _bundle_payload(source: bytes | bytearray | memoryview | Path | str) -> bytes:
    """Read a bundle without accepting an arbitrary archive object."""

    if isinstance(source, bytes):
        payload = source
    elif isinstance(source, (bytearray, memoryview)):
        payload = bytes(source)
    elif isinstance(source, (Path, str)):
        path = Path(source)
        if not path.is_file():
            raise Atlanta6AreaPlanError("BUNDLE_NOT_FOUND")
        payload = path.read_bytes()
    else:
        raise Atlanta6AreaPlanError("BUNDLE_TYPE_NOT_ALLOWED")
    if not payload:
        raise Atlanta6AreaPlanError("BUNDLE_EMPTY")
    return payload


def _strict_csv_rows(payload: bytes, columns: tuple[str, ...], *, artifact: str) -> list[dict[str, str]]:
    """Read one generated CSV and reject any non-builder representation."""

    try:
        text = payload.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text, newline=""), strict=True)
        if tuple(reader.fieldnames or ()) != columns:
            raise Atlanta6AreaPlanError("BUNDLE_ARTIFACT_SCHEMA_INVALID", artifact)
        rows: list[dict[str, str]] = []
        for raw in reader:
            if None in raw or set(raw) != set(columns):
                raise Atlanta6AreaPlanError("BUNDLE_ARTIFACT_ROW_INVALID", artifact)
            rows.append({column: str(raw[column]) for column in columns})
    except Atlanta6AreaPlanError:
        raise
    except (UnicodeDecodeError, csv.Error) as exc:
        raise Atlanta6AreaPlanError("BUNDLE_ARTIFACT_INVALID", artifact) from exc
    # The builder fixes encoding, header order, line endings, and no-op values.
    if _csv_bytes(columns, rows) != payload:
        raise Atlanta6AreaPlanError("BUNDLE_ARTIFACT_NOT_CANONICAL", artifact)
    return rows


def _manifest_payload(payload: bytes) -> dict[str, Any]:
    try:
        decoded = payload.decode("utf-8")
        value = json.loads(decoded)
    except (UnicodeDecodeError, json.JSONDecodeError) as exc:
        raise Atlanta6AreaPlanError("BUNDLE_MANIFEST_INVALID") from exc
    if not isinstance(value, dict):
        raise Atlanta6AreaPlanError("BUNDLE_MANIFEST_INVALID")
    canonical = (json.dumps(value, ensure_ascii=False, indent=2, sort_keys=True) + "\n").encode("utf-8")
    if payload != canonical:
        raise Atlanta6AreaPlanError("BUNDLE_MANIFEST_NOT_CANONICAL")
    return value


def _require_exact_mapping(value: object, expected: Mapping[str, Any], *, code: str) -> dict[str, Any]:
    if not isinstance(value, dict) or value != dict(expected):
        raise Atlanta6AreaPlanError(code)
    return value


def _validated_region(territory: str, region_id: str, region_seq: str, region_name: str) -> None:
    if territory not in ZONE_TO_SEQ:
        raise Atlanta6AreaPlanError("BUNDLE_REGION_INVALID")
    sequence = ZONE_TO_SEQ[territory]
    if (
        region_id != _region_id(territory)
        or region_seq != str(sequence)
        or region_name != _region_name(territory)
    ):
        raise Atlanta6AreaPlanError("BUNDLE_REGION_INVALID")


def _validate_fixed_region_rows(rows: list[dict[str, str]]) -> None:
    if len(rows) != EXPECTED_UNIQUE_POSTALS:
        raise Atlanta6AreaPlanError("BUNDLE_FIXED_REGION_ROW_COUNT_INVALID")
    postals: set[str] = set()
    sequences: set[str] = set()
    for row in rows:
        postal = row["POSTAL_CODE"]
        if not _POSTAL_RE.fullmatch(postal) or postal in postals:
            raise Atlanta6AreaPlanError("BUNDLE_FIXED_REGION_POSTAL_INVALID")
        postals.add(postal)
        if (
            row["STRATEGIC_CITY_NAME"] != STRATEGIC_CITY_NAME
            or row["area_type"] != CANONICAL_AREA_TYPE
        ):
            raise Atlanta6AreaPlanError("BUNDLE_FIXED_REGION_VALUE_INVALID")
        _validated_region(
            row["AREA_NAME"], row["region_id"], row["region_seq"], row["new_region_name"]
        )
        sequences.add(row["region_seq"])
    if sequences != {str(value) for value in ZONE_TO_SEQ.values()}:
        raise Atlanta6AreaPlanError("BUNDLE_FIXED_REGION_COVERAGE_INVALID")


def _validate_boundary_policy_rows(
    rows: list[dict[str, str]],
    *,
    plan_id: str,
    resolutions: Mapping[str, Mapping[str, Any]],
) -> None:
    expected_postals = [
        postal for postal in EXPECTED_AMBIGUOUS_POSTALS if resolutions[postal]["allow_overflow"] is True
    ]
    if len(rows) != len(expected_postals) or [row["POSTAL_CODE"] for row in rows] != expected_postals:
        raise Atlanta6AreaPlanError("BUNDLE_BOUNDARY_POLICY_ROW_ACCOUNTING_INVALID")
    for row, postal in zip(rows, expected_postals):
        decision = resolutions[postal]
        primary = str(decision["primary_region"])
        alternate = str(decision["alternate_region"])
        if (
            row["plan_id"] != plan_id
            or row["STRATEGIC_CITY_NAME"] != STRATEGIC_CITY_NAME
            or row["primary_region_id"] != _region_id(primary)
            or row["primary_region_seq"] != str(ZONE_TO_SEQ[primary])
            or row["alternate_region_id"] != _region_id(alternate)
            or row["alternate_region_seq"] != str(ZONE_TO_SEQ[alternate])
            or row["policy_mode"] != "boundary_primary_alternate_soft"
            or row["penalty_cost"] != str(BOUNDARY_PENALTY_COST)
        ):
            raise Atlanta6AreaPlanError("BUNDLE_BOUNDARY_POLICY_INVALID")


def _validate_technician_policy_rows(rows: list[dict[str, str]], *, plan_id: str) -> None:
    if len(rows) != EXPECTED_TECHNICIAN_ROWS:
        raise Atlanta6AreaPlanError("BUNDLE_TECHNICIAN_ROW_COUNT_INVALID")
    technician_ids = [row["SVC_ENGINEER_CODE"] for row in rows]
    if (
        technician_ids != sorted(technician_ids)
        or len(set(technician_ids)) != EXPECTED_TECHNICIAN_ROWS
        or any(not _TECH_ID_RE.fullmatch(code) for code in technician_ids)
    ):
        raise Atlanta6AreaPlanError("BUNDLE_TECHNICIAN_ID_INVALID")
    for row in rows:
        if (
            row["plan_id"] != plan_id
            or row["STRATEGIC_CITY_NAME"] != STRATEGIC_CITY_NAME
            or row["policy_mode"] != "assigned_region_boundary_spillover"
        ):
            raise Atlanta6AreaPlanError("BUNDLE_TECHNICIAN_POLICY_INVALID")
        _validated_region(
            row["assigned_region_name"].removeprefix(f"{STRATEGIC_CITY_NAME} "),
            row["assigned_region_id"],
            row["assigned_region_seq"],
            row["assigned_region_name"],
        )


def _validate_bundle_manifest(
    manifest: dict[str, Any], artifact_payloads: Mapping[str, bytes]
) -> tuple[str, Mapping[str, Mapping[str, Any]], dict[str, Any]]:
    expected_keys = {
        "schema", "schema_version", "policy_version", "plan_id", "resolution_digest",
        "strategic_city_name", "approval_status", "promotable", "verification_only",
        "lifecycle_stage", "source", "boundary_resolutions", "row_accounting",
        "zone_to_region_seq", "artifacts", "parents", "canonical_artifacts",
        "privacy_classification", "technician_names_redacted",
        "source_technician_master_context", "target_technician_master_context",
    }
    if set(manifest) != expected_keys:
        raise Atlanta6AreaPlanError("BUNDLE_MANIFEST_SCHEMA_INVALID")
    if (
        manifest["schema"] != MANIFEST_SCHEMA
        or manifest["schema_version"] != MANIFEST_SCHEMA
        or manifest["policy_version"] != POLICY_VERSION
        or manifest["strategic_city_name"] != STRATEGIC_CITY_NAME
        or manifest["approval_status"] != "resolved_for_development_verification"
        or manifest["promotable"] is not False
        or manifest["verification_only"] is not True
        or manifest["lifecycle_stage"] != "resolved_candidate"
        or manifest["privacy_classification"] != _PRIVACY_CLASSIFICATION
        or manifest["technician_names_redacted"] is not True
    ):
        raise Atlanta6AreaPlanError("BUNDLE_MANIFEST_VALUE_INVALID")
    _require_exact_mapping(
        manifest["source_technician_master_context"],
        _SOURCE_TECHNICIAN_MASTER_CONTEXT,
        code="BUNDLE_TECHNICIAN_CONTEXT_INVALID",
    )
    _require_exact_mapping(
        manifest["target_technician_master_context"],
        _TARGET_TECHNICIAN_MASTER_CONTEXT,
        code="BUNDLE_TECHNICIAN_CONTEXT_INVALID",
    )
    source = manifest["source"]
    if not isinstance(source, dict) or not re.fullmatch(
        r"[0-9a-f]{64}", str(source.get("sha256", ""))
    ):
        raise Atlanta6AreaPlanError("BUNDLE_SOURCE_LINEAGE_INVALID")
    source_sha256 = str(source["sha256"])
    _require_exact_mapping(
        source,
        _source_metadata(source_sha256),
        code="BUNDLE_SOURCE_LINEAGE_INVALID",
    )
    if manifest["parents"] != [
        {
            "kind": "source_workbook",
            "file_name": SOURCE_FILE_NAME,
            "sha256": source_sha256,
        }
    ]:
        raise Atlanta6AreaPlanError("BUNDLE_SOURCE_LINEAGE_INVALID")
    _require_exact_mapping(manifest["zone_to_region_seq"], ZONE_TO_SEQ, code="BUNDLE_ZONE_MAP_INVALID")
    _require_exact_mapping(
        manifest["canonical_artifacts"], _canonical_artifact_contract(),
        code="BUNDLE_CANONICAL_SCHEMA_INVALID",
    )

    raw_resolutions = manifest["boundary_resolutions"]
    if not isinstance(raw_resolutions, dict):
        raise Atlanta6AreaPlanError("BUNDLE_BOUNDARY_RESOLUTIONS_INVALID")
    try:
        resolutions = _normalize_boundary_resolutions(raw_resolutions)
    except Atlanta6AreaPlanError as exc:
        raise Atlanta6AreaPlanError(f"BUNDLE_{exc.code}") from exc
    if set(raw_resolutions) != set(resolutions) or any(
        raw_resolutions[postal] != dict(resolutions[postal]) for postal in resolutions
    ):
        raise Atlanta6AreaPlanError("BUNDLE_BOUNDARY_RESOLUTIONS_INVALID")
    resolution_digest, plan_id = derive_atlanta_6area_plan_identity(
        source_sha256, resolutions
    )
    if manifest["resolution_digest"] != resolution_digest or manifest["plan_id"] != plan_id:
        raise Atlanta6AreaPlanError("BUNDLE_PLAN_ID_INVALID")

    overflow_count = sum(1 for item in resolutions.values() if item["allow_overflow"] is True)
    expected_accounting = {
        "membership": {
            "input": EXPECTED_MEMBERSHIP_ROWS,
            "accepted": EXPECTED_MEMBERSHIP_ROWS,
            "rejected": 0,
            "unique_postals": EXPECTED_UNIQUE_POSTALS,
            "ambiguous_postals": len(EXPECTED_AMBIGUOUS_POSTALS),
        },
        "technician": {
            "input": EXPECTED_TECHNICIAN_ROWS,
            "accepted": EXPECTED_TECHNICIAN_ROWS,
            "rejected": 0,
        },
    }
    _require_exact_mapping(
        manifest["row_accounting"], expected_accounting, code="BUNDLE_ROW_ACCOUNTING_INVALID"
    )
    artifact_rows = {
        FIXED_REGION_FILENAME: EXPECTED_UNIQUE_POSTALS,
        BOUNDARY_POLICY_FILENAME: overflow_count,
        TECHNICIAN_POLICY_FILENAME: EXPECTED_TECHNICIAN_ROWS,
    }
    artifact_manifest = manifest["artifacts"]
    if not isinstance(artifact_manifest, dict) or set(artifact_manifest) != set(artifact_rows):
        raise Atlanta6AreaPlanError("BUNDLE_ARTIFACT_MANIFEST_INVALID")
    for name, row_count in artifact_rows.items():
        payload = artifact_payloads[name]
        _require_exact_mapping(
            artifact_manifest[name],
            _artifact_metadata(payload, row_count),
            code="BUNDLE_ARTIFACT_HASH_INVALID",
        )
    summary = {
        "plan_id": plan_id,
        "strategic_city_name": STRATEGIC_CITY_NAME,
        "source_sha256": source_sha256,
        "membership_input_rows": EXPECTED_MEMBERSHIP_ROWS,
        "membership_accepted_rows": EXPECTED_MEMBERSHIP_ROWS,
        "membership_rejected_rows": 0,
        "unique_postal_count": EXPECTED_UNIQUE_POSTALS,
        "ambiguous_postal_count": len(EXPECTED_AMBIGUOUS_POSTALS),
        "ambiguous_postals": list(EXPECTED_AMBIGUOUS_POSTALS),
        "technician_input_rows": EXPECTED_TECHNICIAN_ROWS,
        "technician_accepted_rows": EXPECTED_TECHNICIAN_ROWS,
        "technician_rejected_rows": 0,
        "fixed_region_rows": EXPECTED_UNIQUE_POSTALS,
        "canonical_region_count": len(ZONE_TO_SEQ),
        "canonical_postal_count": EXPECTED_UNIQUE_POSTALS,
        "canonical_technician_count": EXPECTED_TECHNICIAN_ROWS,
        "canonical_boundary_resolution_count": len(resolutions),
        "canonical_boundary_policy_count": overflow_count,
        "boundary_resolution_rows": len(resolutions),
        "boundary_policy_rows": overflow_count,
        "technician_policy_rows": EXPECTED_TECHNICIAN_ROWS,
        "boundary_overflow_enabled_count": overflow_count,
        "boundary_penalty_cost": BOUNDARY_PENALTY_COST,
    }
    return plan_id, resolutions, summary


def validate_atlanta_6area_bundle(
    source: bytes | bytearray | memoryview | Path | str,
) -> Atlanta6AreaBundle:
    """Validate the exact deterministic development candidate bundle.

    Validation is deliberately self-contained and read-only: it verifies the
    archive layout, canonical bytes, source lineage, all manifest hashes,
    schemas, duplicate semantics, and row accounting.  It never extracts to
    disk, stages a file, or writes a database.
    """

    bundle_bytes = _bundle_payload(source)
    expected_names = {
        FIXED_REGION_FILENAME,
        BOUNDARY_POLICY_FILENAME,
        TECHNICIAN_POLICY_FILENAME,
        MANIFEST_FILENAME,
    }
    try:
        with zipfile.ZipFile(io.BytesIO(bundle_bytes), "r") as archive:
            infos = archive.infolist()
            names = [info.filename for info in infos]
            if len(names) != len(set(names)) or set(names) != expected_names:
                raise Atlanta6AreaPlanError("BUNDLE_ARCHIVE_LAYOUT_INVALID")
            if any(info.is_dir() or info.file_size > 8 * 1024 * 1024 for info in infos):
                raise Atlanta6AreaPlanError("BUNDLE_ARCHIVE_LAYOUT_INVALID")
            if sum(info.file_size for info in infos) > 16 * 1024 * 1024:
                raise Atlanta6AreaPlanError("BUNDLE_ARCHIVE_TOO_LARGE")
            if archive.testzip() is not None:
                raise Atlanta6AreaPlanError("BUNDLE_ARCHIVE_INVALID")
            artifact_payloads = {name: archive.read(name) for name in names}
    except Atlanta6AreaPlanError:
        raise
    except (OSError, RuntimeError, zipfile.BadZipFile) as exc:
        raise Atlanta6AreaPlanError("BUNDLE_ARCHIVE_INVALID") from exc

    manifest = _manifest_payload(artifact_payloads[MANIFEST_FILENAME])
    plan_id, resolutions, summary = _validate_bundle_manifest(manifest, artifact_payloads)
    fixed_rows = _strict_csv_rows(
        artifact_payloads[FIXED_REGION_FILENAME], _FIXED_REGION_COLUMNS, artifact=FIXED_REGION_FILENAME
    )
    boundary_rows = _strict_csv_rows(
        artifact_payloads[BOUNDARY_POLICY_FILENAME], _BOUNDARY_POLICY_COLUMNS, artifact=BOUNDARY_POLICY_FILENAME
    )
    technician_rows = _strict_csv_rows(
        artifact_payloads[TECHNICIAN_POLICY_FILENAME], _TECHNICIAN_POLICY_COLUMNS, artifact=TECHNICIAN_POLICY_FILENAME
    )
    _validate_fixed_region_rows(fixed_rows)
    _validate_boundary_policy_rows(boundary_rows, plan_id=plan_id, resolutions=resolutions)
    _validate_technician_policy_rows(technician_rows, plan_id=plan_id)
    if _deterministic_zip(artifact_payloads) != bundle_bytes:
        raise Atlanta6AreaPlanError("BUNDLE_NOT_DETERMINISTIC")
    return Atlanta6AreaBundle(
        plan_id=plan_id,
        promotable=False,
        approval_status="resolved_for_development_verification",
        summary=MappingProxyType(summary),
        manifest=MappingProxyType(manifest),
        artifacts=MappingProxyType(dict(artifact_payloads)),
        bundle_bytes=bundle_bytes,
    )


__all__ = [
    "AREA_SHEET_NAME",
    "Atlanta6AreaBundle",
    "Atlanta6AreaPlanError",
    "Atlanta6AreaPreview",
    "BOUNDARY_PENALTY_COST",
    "BOUNDARY_POLICY_FILENAME",
    "BOUNDARY_POLICY_SCHEMA",
    "CANONICAL_AREA_TYPE",
    "EXPECTED_AMBIGUOUS_POSTALS",
    "EXPECTED_SOURCE_SHA256",
    "EXPECTED_TECHNICIAN_ROWS",
    "FIXED_REGION_FILENAME",
    "FIXED_REGION_SCHEMA",
    "MANIFEST_FILENAME",
    "PLAN_ID",
    "POLICY_VERSION",
    "STRATEGIC_CITY_NAME",
    "TECHNICIAN_POLICY_FILENAME",
    "TECHNICIAN_POLICY_SCHEMA",
    "TECHNICIAN_SHEET_NAME",
    "ZONE_TO_SEQ",
    "build_atlanta_6area_bundle",
    "build_atlanta_6area_template_bytes",
    "derive_atlanta_6area_plan_identity",
    "get_atlanta_6area_template_bytes",
    "get_atlanta_6area_workbook_template",
    "is_allowed_atlanta_6area_plan_id",
    "parse_atlanta_6area_workbook",
    "preview_atlanta_6area_plan",
    "validate_atlanta_6area_bundle",
]
